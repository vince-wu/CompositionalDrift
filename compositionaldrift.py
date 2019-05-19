#All neccesary imports
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import random
import math
import json
import os.path
matplotlib.use('TkAgg')
from numpy import arange, sin, pi
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
from matplotlib import style
# implement the default mpl key bindings
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure
import sys
if sys.version_info[0] < 3:
    import Tkinter as Tk
    from Tkinter import ttk
    import Tkinter.filedialog
    import Tkinter.messagebox
    print(sys.version_info[0], "hi")
else:
    import tkinter as Tk
    from tkinter import ttk
    from tkinter.colorchooser import *
    import tkinter.filedialog
    import tkinter.messagebox
    print(sys.version_info[0])
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw
from openpyxl.drawing.image import Image as xlImage
"""
    Created by Vincent Wu on 8/17/16:
    This program uses the Mayo-Lewis equation and Monte Carlo method to simulate copolymer growth, and
    represents results both graphically and visually
    """
 #test

"---------------------------------------------------------------------------------------------------------------------------------"
"***STATIC FINAL VARIABLES***"
"---------------------------------------------------------------------------------------------------------------------------------"

MONOMER_CAP = 5000000
NUM_UNIQUE_MONOMERS = 2
NUM_SIMULATIONS = 1
NUM_POLYMERS_SHOW = 8
GRAPH_TYPE = 1
TOTAL_STARTING_MONOMERS = 200000
DP = 100
HYDROPHOBICITY = 0
SETTING = 1
LOAD_SUCCESSFUL = False
GRAPH1_TYPE = "Monomer Occurrences"
GRAPH2_TYPE = "Monomer Usage"
HISTOGRAM1_MONOMER = 1
HISTOGRAM2_MONOMER = 2
HISTOGRAM_LIMIT = 1
PENULTIMATE = 0
DYAD = 0
LEGEND = 1
ALIAS = 0
SETINITIAL = 0
CONVERSION = 100
MAINTAIN = 0
STYLE = "bmh"
COLOR1 = '#4D4D4D'
COLOR2 = '#5DA5DA'
COLOR3 = '#F15854'
COLOR4 = '#DECF3F'
COLOR5 = '#60BD68'
COLOR6 = '#F17CB0'	
COLOR7 = '#B276B2'
COLOR8 = '#FAA43A'
DCOLOR1 = '#353535'
DCOLOR2 = '#4a84ae'
DCOLOR3 = '#c04643'
DCOLOR4 = '#b1a532'
DCOLOR5 = '#4c9753'
DCOLOR6 = '#c0638c'
DCOLOR7 = '#8e5e8e'
DCOLOR8 = '#c8832e'
COLORARRAY = [COLOR1, COLOR2, COLOR3, COLOR4, COLOR5, COLOR6, COLOR7, COLOR8]
DCOLORARRAY = [DCOLOR1, DCOLOR2, DCOLOR3, DCOLOR4, DCOLOR5, DCOLOR6, DCOLOR7, DCOLOR8]
DYADCOLORARRAY = [COLOR1, COLOR2, COLOR3, COLOR4, COLOR5, COLOR6, COLOR7, COLOR8]
VERSION = "v1.9.0.1"
CONFIGS = [["Number of Unique Monomers", 1], ["Number of Simulations", 1],
 ["Number of Polymers to Show", 1], 
 ["Graph Monomer Occurence", 1], ["Monomer Pool Size", 1], ["Monomers to RAFT Ratio", 1], 
 ["Default Setting", 1], ["Monomer Cap", 1], ["Graph 1 Type", 1], ["Graph 2 Type", 1], ["Histogram 1 Monomer", 1], ["Histogram 2 Monomer", 1],
 ["Percentage to Analyze for Histogram", 0], ["Penultimate", 1], ["Dyad", 1], ["Style", 2], ["Legend", 1], ["Percent Conversion", 0],
 ["Color1", 2], ["Color2", 2], ["Color3", 2], ["Color4", 2], ["Color5", 2], ["Color6", 2], ["Color7", 2], ["Color8", 2], ["Maintain", 1], ["Hydrophobicity", 0]]

"----------------------------------------------------------------------------------------------------------------------------------------------------------------------"
"***CONFIG FILE READ AND WRITE***"
"----------------------------------------------------------------------------------------------------------------------------------------------------------------------"

#generates a config file if needed
def generateConfigFile():
	if not os.path.exists("config.txt"):
		file = open("config.txt", "w")
		file.write("Number of Unique Monomers = 2 \nNumber of Simulations = 1 \nNumber of Polymers to Show = 8 \n")
		file.write("Graph 1 Type = 0 \nGraph 2 Type = 1 \n")
		file.write("Histogram 1 Monomer = 1 \nHistogram 2 Monomer = 2 \nPercentage to Analyze for Histogram = 1 \n")
		file.write("Monomer Pool Size = 200000 \nMonomers to RAFT Ratio = 100 \nHydrophobicity = 0 \nDefault Setting = 1 \nMonomer Cap = 5000000 \nPenultimate = 0 \n")
		file.write("Dyad = 0 \nStyle = classic \nLegend = 1 \nPercent Conversion = 100 \nMaintain = 0 \n")
		file.write("Color1 = #4D4D4D \nColor2 = #5DA5DA \nColor3 = #F15854 \nColor4 = #DECF3F \nColor5 = #60BD68 \nColor6 = #F17CB0 \n")
		file.write("Color7 = #B276B2 \nColor8 = #FAA43A \n")
		file.close()
#generates a save file if needed
def generateSaveFile():
	if not os.path.exists("state1.txt"):
		file = open("state1.txt", "w")
		file.write("# use the '#' character followed by a space to denote a line you want to use as a comment, \n")
		file.write("# it will not be read by the config file and won't mess up your settings. \n# Sample Save File \n")
		file.write("Number of Unique Monomers = 4 \nMonomer 1 Ratio = 50 \nMonomer 2 Ratio = 25 \nMonomer 3 Ratio = 20 \nMonomer 4 Ratio = 5 \n") 
		file.write("1-1 = 0.89 \n1-2 = 1 \n1-3 = 1 \n1-4 = 1 \n2-1 = 1 \n2-2 = 1.1 \n2-3 = 1.1 \n2-4 = 1.1 \n3-1 = 1 \n3-2 = 1.1 \n3-3 = 1.1 \n3-4 = 1.1 \n")
		file.write("4-1 = 1 \n4-2 = 1.1 \n4-3 = 1.1 \n4-4 = 1.1 \nend")
		file.close()
#reads config before initScreen for early variables
def earlyConfigRead():
	#error checking to see if config file exists
	if not os.path.exists("config.txt"):
		errorMessage("config.txt does not exist!", 220)
		return
	file = open("config.txt", "r")
	#variable to keep track of number of invalid config lines
	validLines = 0 
	notParsed = True
	for line in file:
		line = line.strip()
		intermediate = line.split("=")
		if len(intermediate) == 2:
			global NUM_UNIQUE_MONOMERS
			global SETTING
			global PENULTIMATE
			try:
				if intermediate[0].strip() == "Number of Unique Monomers" and notParsed:
					NUM_UNIQUE_MONOMERS = int(intermediate[1].strip())
					notParsed = False
					validLines += 1
				if intermediate[0].strip() == "Default Setting":
					SETTING = int(intermediate[1].strip())
					validLines += 1
				if intermediate[0].strip() == "Penultimate":
					PENULTIMATE= int(intermediate[1].strip())
					validLines += 1
					print("read penultimate!")
			except ValueError:
				continue
	print("earlyConfigRead valid lines: %s" %validLines)
	file.close()
#reads the config file and sets static variables based on config
def readConfigFile():
	#error checking to see if config file exists
	if not os.path.exists("config.txt"):
		errorMessage("config.txt does not exist!", 220)
		return
	file = open("config.txt", "r")
	#parses each line in the file. If it is a valid configuration line, it sets that correspondin static variable accordingly
	#variable to keep track of number of invalid config lines
	invalidLines = 0 
	#variable to keep track of whether or not line is a setting config line
	settingConfig = False
	#variable to keep track of whether to parse setting config
	parseSetting = False
	#variable to keep track of number of monomers
	global numMonomers
	numMonomers = 0
	#global array to keep track of monomer ratios, initalized as an array of -1
	global RATIO_ARRAY
	#global array to keep track monomer coefficients, initialized of an array of arrays of -1
	global COEFF_ARRAY
	#global variable to shwo whether or not a setting loaded successfully
	global LOAD_SUCCESSFUL
	#global variable to kep track of last setting number
	global LAST_SETTING
	LAST_SETTING = 0
	LOAD_SUCCESSFUL = False
	for line in file:
		line = line.strip()
		#If line is end, read next lines as regular config line
		#handling for setting config line
		try:
			#gets the config header
			configType = (line.split("="))[0].strip()
			#gets the config value
			configStringValue = (line.split("="))[1].strip()
		except:
			invalidLines += 1
			continue
		if (not setConfigVariable(configType, configStringValue)):
			invalidLines += 1
			continue
	LOAD_SUCCESSFUL = True
	file.close()
	print("Number of invalid config lines: ", invalidLines)
def readSaveFile(stateNumber):
	file = open("state%i.txt" %(stateNumber), "r")
	#parses each line in the file. If it is a valid configuration line, it sets that correspondin static variable accordingly
	#variable to keep track of number of invalid config lines
	invalidLines = 0 
	#variable to keep track of whether or not line is a setting config line
	settingConfig = False
	#variable to keep track of whether to parse setting config
	parseSetting = False
	#variable to keep track of number of monomers
	global numMonomers
	numMonomers = 0
	#global array to keep track of monomer ratios, initalized as an array of -1
	global RATIO_ARRAY
	#global array to keep track monomer coefficients, initialized of an array of arrays of -1
	global COEFF_ARRAY
	#global variable to shwo whether or not a setting loaded successfully
	global LOAD_SUCCESSFUL
	#global variable to kep track of last setting number
	global LAST_SETTING
	#variable to keep track of model (penultimate vs standard)
	model = "standard"
	LAST_SETTING = 0
	LOAD_SUCCESSFUL = False
	for line in file:
		line = line.strip()
		lineArray = line.split(" ")
		if lineArray[0] == '#':
			continue
		if line == "end":
			#test to see that the settings are all valid and can be loaded properly
			#standard case:
			if model == "standard":
				try:
					#test to see that ratios for all monomers are given
					for monomerID in numMonomerArray:
						assert(numMonomerArray[monomerID -1] == monomerID)
					#test to see that all ratios are valid (> 0)
					for ratio in RATIO_ARRAY:
						assert(ratio > 0)
					for column in COEFF_ARRAY:
						for coeff in column:
							assert(coeff >= 0)
				except AssertionError:
					parseSetting = False
					LOAD_SUCCESSFUL = False
					continue
				except IndexError:
					parseSetting = False
					LOAD_SUCCESSFUL = False
				LOAD_SUCCESSFUL = True
			if model == "penultimate":
				LOAD_SUCCESSFUL = True
		#setting correct models
		if line == "Standard":
			model = "standard"
			global PENULTIMATE
			PENULTIMATE = False
		if line == "Penultimate":
			model = "penultimate"
			#setting global variable
			global PENULTIMATE
			PENULTIMATE = True
			#readjust structure of nested array to penultimate structure
			COEFF_ARRAY = [[[-1 for i in range(numMonomers)] for j in range(numMonomers)] for k in range(numMonomers)]
		if len(lineArray) == 6:
			try:
				assert(lineArray[0] == "Number")
				assert(lineArray[1] == "of")
				assert(lineArray[2] == "Unique")
				assert(lineArray[3] == "Monomers")
				assert(lineArray[4] == "=")
				numMonomers = int(lineArray[5])
				RATIO_ARRAY = [-1] * numMonomers
				#array to keep track of all neccesary monomers
				if numMonomers == 2:
					numMonomerArray = [0] * numMonomers
					COEFF_ARRAY = [-1] * numMonomers
					COEFF_ARRAY = [[-1 for i in range(numMonomers)] for j in range(numMonomers)]
				else:
					numMonomerArray = [0] * numMonomers
					COEFF_ARRAY = [-1] * numMonomers
					COEFF_ARRAY = [[] for j in range(numMonomers)]
			except AssertionError:
				invalidLines += 1
			except ValueError:
				invalidLines += 1
			continue
		if len(lineArray) == 5:
			#checking that line is correct syntax
			try:
				assert(lineArray[0] == "Monomer")
				assert(lineArray[2] == "Ratio")
				assert(lineArray[3] == "=")
				#adding monomer number to numMonomerArray
				numMonomerArray[int(lineArray[1]) - 1] = int(lineArray[1])
				#add monomer ratio to RATIO_ARRAY
				RATIO_ARRAY[int(lineArray[1]) - 1] = float(lineArray[4])
			except AssertionError:
				invalidLines += 1
				continue
			except ValueError:
				invalidLines += 1
				continue 
			except IndexError:
				invalidLines += 1
				continue
		if len(lineArray) == 3:
			#add the coefficient to the correct place in COEFF_ARRAY
			#standard case
			if model == "standard":
				try:
					assert(len(lineArray[0]) == 3)
					assert(lineArray[1] == "=")
					coeffTag = lineArray[0]
					assert(coeffTag[1] == "-")
					column = int(coeffTag[0]) - 1
					if numMonomers == 2:
						row = int(coeffTag[2]) - 1
						COEFF_ARRAY[column][row] = float(lineArray[2])
					else:
						COEFF_ARRAY[column].append(float(lineArray[2]))
				except AssertionError:
					invalidLines += 1
					continue
				except ValueError:
					invalidLines += 1
					continue
				except IndexError:
					invalidLines += 1
					continue
			#penultimate case
			if model == "penultimate":
				try:
					#check to see if number of character is 5 (1-2-3 has 5 characters, as should all penultimate coefficients)
					assert(len(lineArray[0]) == 5)
					#check to see if second character is rquals sign (1-2-1 = 5, the second character should always be '=')
					assert(lineArray[1] == "=")
					#getting the place where the coefficient should go
					coeffTag = lineArray[0]
					#asserting that there are 2 dashes in the right place
					assert(coeffTag[1] == "-")
					assert(coeffTag[3] == "-")
					#use the coeffTag to determing correct placement in nested array
					penultimate = int(coeffTag[0]) - 1
					ultimate = int(coeffTag[2]) - 1
					nextMonomer = int(coeffTag[4]) - 1
					print("penultimate: ", penultimate)
					print("ultimate: ", ultimate)
					print("nextMonomer: ", nextMonomer)
					#adding value of coeff to correct place in nested array
					COEFF_ARRAY[nextMonomer][ultimate][penultimate] = float(lineArray[2])
				except AssertionError:
					invalidLines += 1
					continue
				except ValueError:
					invalidLines += 1
					continue
				except IndexError:
					invalidLines += 1
					continue
		else:
			invalidLines += 1
			continue
		try:
			#gets the config header
			configType = (line.split("="))[0].strip()
			#gets the config value
			configStringValue = (line.split("="))[1].strip()
		except:
			invalidLines += 1
			continue
		if (not setConfigVariable(configType, configStringValue)):
			invalidLines += 1
			continue
	print("ran loadSave")
	print("unique monomers: ", NUM_UNIQUE_MONOMERS)
	print("ratio array: ", RATIO_ARRAY)
	print("coeff array: ", COEFF_ARRAY)
	file.close()
#helper method to set the static variable. Returns 1 if successful, 0 if not
def setConfigVariable(configType, configStringValue):
	for config in CONFIGS:
		if config[0] == configType:
			try:
				#handling for variable types
				if config[1] == 0:
					setConfigVariableHelper(configType, float(configStringValue))
				elif config[1] == 1:
					setConfigVariableHelper(configType, int(configStringValue))
				elif config[1] == 2:
					#print("configstring ", configStringValue)
					setConfigVariableHelper(configType, configStringValue)
				print(configType, configStringValue)
				return 1
			except AssertionError:
				return 0
	return 0
#A helper function mapping name to static variables
def setConfigVariableHelper(configType, configValue):
	if configType == "Number of Unique Monomers":
		global NUM_UNIQUE_MONOMERS
		NUM_UNIQUE_MONOMERS = configValue
	elif configType == "Number of Simulations":
		global NUM_SIMULATIONS
		NUM_SIMULATIONS = configValue
	elif configType == "Number of Polymers to Show":
		global NUM_POLYMERS_SHOW
		NUM_POLYMERS_SHOW = configValue	
	elif configType == "Monomer Pool Size":
		global TOTAL_STARTING_MONOMERS
		TOTAL_STARTING_MONOMERS = configValue
	elif configType == "Monomers to RAFT Ratio":
		global DP
		DP = configValue
	elif configType == "Default Setting":
		pass
	elif configType == "Monomer Cap":
		global MONOMER_CAP
		MONOMER_CAP = configValue
	elif configType == "Graph 1 Type":
		global GRAPH1_TYPE
		if configValue == 0:
			GRAPH1_TYPE = "Monomer Occurrences"
		elif configValue == 1:
			GRAPH1_TYPE = "Monomer Usage"
		elif configValue == 2:
			GRAPH1_TYPE = "Monomer Separation"
		elif configValue == 3:
			GRAPH1_TYPE = "None"
		elif configValue == 4:
			GRAPH1_TYPE = "DP Distribution"
		elif configValue == 5:
			GRAPH1_TYPE = "Hydrophobic Blocks"
		elif configValue == 6:
			GRAPH1_TYPE ="Hydrophilic Blocks"
		else:
			assert False
	elif configType == "Graph 2 Type":
		global GRAPH2_TYPE
		if configValue == 0:
			GRAPH2_TYPE = "Monomer Occurrences"
		elif configValue == 1:
			GRAPH2_TYPE = "Monomer Usage"
		elif configValue == 2:
			GRAPH2_TYPE = "Monomer Separation"
		elif configValue == 3:
			GRAPH2_TYPE = "None"
		elif configValue == 4:
			GRAPH2_TYPE = "DP Distribution"
		elif configValue == 5:
			GRAPH2_TYPE = "Hydrophobic Blocks"
		elif configValue == 6:
			GRAPH2_TYPE = "Hydrophilic Blocks"
		else:
			assert False
	elif configType == "Histogram 1 Monomer":
		global HISTOGRAM1_MONOMER
		HISTOGRAM1_MONOMER = configValue
	elif configType == "Histogram 2 Monomer":
		global HISTOGRAM2_MONOMER
		HISTOGRAM2_MONOMER = configValue
	elif configType == "Percentage to Analyze for Histogram":
		global HISTOGRAM_LIMIT
		HISTOGRAM_LIMIT = configValue
	elif configType == "Penultimate":
		global PENULTIMATE
		PENULTIMATE = configValue
	elif configType == "Dyad":
		global DYAD
		DYAD = configValue
	elif configType == "Style":
		global STYLE
		STYLE = configValue
	elif configType == "Legend":
		global LEGEND
		LEGEND = configValue
	elif configType == "Percent Conversion":
		global CONVERSION
		CONVERSION = configValue
	elif configType == "Color1":
		global COLOR1
		COLOR1 = configValue
	elif configType == "Color2":
		global COLOR2
		COLOR2 = configValue
	elif configType == "Color3":
		global COLOR3
		COLOR3 = configValue
	elif configType == "Color4":
		global COLOR4
		COLOR4 = configValue
	elif configType == "Color5":
		global COLOR5
		COLOR5 = configValue
	elif configType == "Color6":
		global COLOR6
		COLOR6 = configValue
	elif configType == "Color7":
		global COLOR7
		COLOR7 = configValue
	elif configType == "Color8":
		global COLOR8
		COLOR8 = configValue
	elif configType == "Maintain":
		global MAINTAIN
		MAINTAIN = configValue
	elif configType == "Hydrophobicity":
		global HYDROPHOBICITY
		HYDROPHOBICITY = configValue
	else:
		assert False, "shouldn't get here"	

"------------------------------------------------------------------------------------------------------------------------------"
"***MAIN APPLICATION CLASS***"
"-------------------------------------------------------------------------------------------------------------------------------"

class Application(ttk.Frame):
	def __init__(self, master = None):
		ttk.Frame.__init__(self, master)
		self.pack()
		self.initialize()
	#Initialization
	def initialize(self):
		try:
			#create config file
			generateConfigFile()
		except:
			errorMessage("Unable to generate config file!", 220)
		try:
			#early read on config file
			earlyConfigRead()
		except:
			errorMessage("Unable to read config file!", 220)
		try:
			#create sample save file
			generateSaveFile()
		except:
			errorMessage("Unable to generate save file!", 220)
		#Creates the init screen
		self.initScreen()
		#Creates Input Widgets
		self.createInputWidgets()
		#Creates dummy visualization Frame

	"------------------------------------------------------------------------------------------------------------------------------"
	"***CREATE THE INITIAL SCREEN, PART 1/2***"
	"------------------------------------------------------------------------------------------------------------------------------"
		#Frame and contents for opening screen
	def initScreen(self):
		root.maxsize(width = 330, height = 270)
		self.initFrame = ttk.Frame(master = root)
		self.initFrame.pack(side = Tk.TOP, fill = Tk.X, expand = 0, padx = 3, pady = 5)
		message1 = Tk.Message(master = self.initFrame, width = 250, font = ("Times New Roman", 10, "bold"),
		 text = "Compositional Drift Simulator %s" % VERSION)
		message1.pack(side = Tk.TOP, padx = 0, pady = 0)
		message2 = Tk.Message(master = self.initFrame, width = 300,
			text = "Author: Vincent Wu (vincent.wu@berkeley.edu)")
		message2.pack(side = Tk.TOP)
		message3 = Tk.Message(master = self.initFrame, width = 300,
			text = "Welcome to Compositional Drift Simulator! This program uses the Mayo Lewis Equation and the" 
			" Monte Carlo method to simulate the growth of a copolymer chain."
			" Please input your desired number of unique monomers or choose a default setting.")
		message3.pack(side = Tk.TOP)
		root.resizable(width=True, height=True)
		root.sizefrom(who = "program")

	"-------------------------------------------------------------------------------------------------------------------------"
	"***CREATE THE INTIAL INPUT SCREEN, PART 2/2***"
	"-------------------------------------------------------------------------------------------------------------------------"		
	
	#Creates input widgets
	def createInputWidgets(self):
		self.canvasExists = False
		self.destroyHide = False
		self.dyadTkVar = Tk.IntVar()
		self.dyadTkVar.set(DYAD)
		self.legendTkVar = Tk.IntVar()
		self.legendTkVar.set(LEGEND)
		self.aliasList = []
		self.simulateLocked = False
		self.initialSetTkVar = Tk.StringVar()
		self.initialSetTkVar.set("Weighted")
		self.conversionTkVar = Tk.DoubleVar()
		self.conversionTkVar.set(CONVERSION)
		self.aliasList = []
		self.subplot1Exists = False
		self.subplot2Exists = False
		self.visualizationFrameExists = False
		self.compFrameExists = False
		self.plotted = False
		def loadSettings(self):
			global useLoadedSettings
			useLoadedSettings = True
			global SETTING
			try: 
				SETTING = int(self.settingTkVar.get())
				print("Setting", SETTING)
				assert int(SETTING) >= 0
			except:
				errorMessage("Please input valid parameters!", 220)
				return
			#error checking to see if config file exists
			if not os.path.exists("state%i.txt" %(SETTING)):
				errorMessage("state%i.txt does not exist!" %(SETTING), 220)
				return
			#reads config file
			try:
				readConfigFile()
				readSaveFile(SETTING)
			except :
				pass
			if LOAD_SUCCESSFUL:
				print("Load successful!")
				self.createMoreInputs()
			else:
				errorMessage("Unable to load config settings! Please fix config file.", 300)
		#Confirms number of monomers, creates more input widgets
		def enter(self):
			#errorMessage("Please vemno Vincent Wu @shinyxspoon $15 for continued use of this program.", 500)
			#return
			#read config file
			try:
				readConfigFile()
			except:
				pass
			if LOAD_SUCCESSFUL:

				print("Load Successful!")
			else:
				errorMessage("Unable to load settings! Please fix config file.", 300)
			#nonlocal numMonomers 
			self.numMonomers = int(self.monomerCountTkVar.get())
			try:
				#asserting that input is in correct range
				assert self.numMonomers < 8
				assert self.monomerCountTkVar.get() > 0
			except:
				errorMessage("Please input valid paramters!", 220)
			global useLoadedSettings
			useLoadedSettings = False
			#print(numMonomers) # debugging purposes
			self.createMoreInputs() 
		#Clears unneccesary widgets and creates more input widgets; called by enter()
		def _quit():
			root.quit()
			root.destroy()
		
		#The parent LabelFrame for all Input Widgets
		
		self.inputFrame = ttk.Frame(master = root, borderwidth = 1, relief = "ridge")
		self.inputFrame.pack(side = Tk.BOTTOM, fill = Tk.X, expand = 0, padx = 7, pady = 7)
		#A frame for column 1 of inputs
		self.columnFrame = ttk.Frame(master = self.inputFrame)
		self.columnFrame.pack(side = Tk.LEFT, padx = 5, pady = 5)
		#A frame for row 1 of inputs
		self.rowFrame1 = ttk.Frame(master = self.columnFrame)
		self.rowFrame1.pack(side = Tk.TOP, padx = 5, pady = 0)
		#A frame for row 2 of inputs
		self.rowFrame2 = ttk.Frame(master = self.columnFrame)
		self.rowFrame2.pack(side = Tk.TOP, padx = 5, pady = 5)
		#Label for monomerCount spinbox
		self.monomerCountLabel = ttk.Label(master = self.rowFrame1, text = "Number of Unique Monomers:")
		self.monomerCountLabel.pack(side = Tk.LEFT, padx = 0, pady = 0)
		#tkvar for monomercount
		self.monomerCountTkVar = Tk.IntVar()
		#MonomerCount spinbox
		self.monomerCountSpinbox = Tk.Spinbox(master = self.rowFrame1, from_ = 2, to = 7, width = 2, textvariable = self.monomerCountTkVar)
		self.monomerCountSpinbox.pack(side = Tk.LEFT, padx = 5, pady = 0)
		#sets monomerCOuntTkVar to default setting
		self.monomerCountTkVar.set(NUM_UNIQUE_MONOMERS)
		#countConfirm Button
		self.countConfirmButton = ttk.Button(master = self.rowFrame1, text = "Enter",
		command = lambda:enter(self), width = 9)
		self.countConfirmButton.pack(side = Tk.LEFT, padx = 5, pady = 5)
		#Label for msetting spinbox
		self.monomerCountLabel = ttk.Label(master = self.rowFrame2, text = "Setting Number to Use:")
		self.monomerCountLabel.pack(side = Tk.LEFT, padx = 0, pady = 0)
		#tkvar for setting
		self.settingTkVar = Tk.IntVar()
		#settingCount spinbox
		self.settingSpinbox = Tk.Spinbox(master = self.rowFrame2, from_ = 1, to = 1000, width = 2, textvariable = self.settingTkVar)
		self.settingSpinbox.pack(side = Tk.LEFT, padx = 5, pady = 0)
		#sets settingTkVar to 1
		global SETTING
		self.settingTkVar.set(SETTING)
		#use loaded inputs button
		self.loadButton = ttk.Button(master = self.rowFrame2, text = "Load from Settings",
		command = lambda:loadSettings(self), width = 18)
		self.loadButton.pack(side = Tk.LEFT, padx = 5, pady = 5)
		#frame for penultimate model choice, 3 row of inputs
		self.penultimateFrame = ttk.Frame(master = self.columnFrame)
		self.penultimateFrame.pack(side = Tk.TOP, padx = 5, pady = 0)
		#Label for penultimate checkbox
		self.penultimateLabel = ttk.Label(master = self.penultimateFrame, text = "Use Penultimate Model:")
		self.penultimateLabel.pack(side = Tk.LEFT, padx = 0, pady = 0)
		self.penultimateCheckButton = ttk.Checkbutton(master = self.penultimateFrame, text = None)
		self.penultimateCheckButton.pack(side = Tk.LEFT, padx = 5)
		#Setting penultimateTkVar to PENULTIMATE global variable
		self.penultimateTkVar = Tk.IntVar()
		self.penultimateCheckButton["variable"] = self.penultimateTkVar
		self.penultimateTkVar.set(PENULTIMATE)
		#rightmost separator
		#self.sideSep2 = ttk.Separator(master = self.inputFrame, orient = Tk.VERTICAL)
		#self.sideSep2.pack(side = Tk.LEFT, fill = Tk.BOTH)
		#bottom separator
		#self.bottomSep = ttk.Separator(master = self.inputFrame, orient = Tk.HORIZONTAL)
		#self.bottomSep.pack(side = Tk.BOTTOM, fill = Tk.BOTH)
	#Creates more input widgets based on numMonomers


	"-----------------------------------------------------------------------------------------------------------------------------"
	"***DESTROY UNECCESARY WIDGETS***"
	"-----------------------------------------------------------------------------------------------------------------------------"
	
	#Destroys unneccesary widgets
	def destroyWidgets(self):
		self.inputFrame.destroy()
		if self.visualizationFrameExists:
			self.visualizationFrame.destroy()
		self.visualizationFrameExists = False
		if self.canvasExists:
			self.canvas.get_tk_widget().destroy()
			self.toolbar.destroy()
		if self.compFrameExists:
			self.compFrame.destroy()
			self.compFrameExists = False
		self.canvasExists = False
		self.destroyHide = False
	
	"-----------------------------------------------------------------------------------------------------------------------------------"
	"***CREATES SECOND ROUND OF INPUTS AFTER 'ENTER' OR 'LOAD' IS PRESSED"
	"-----------------------------------------------------------------------------------------------------------------------------------"
	
	def createMoreInputs(self):
		root.maxsize(width = 10000, height = 10000)
		root.resizable(width = True, height = True)
		root.sizefrom(who = "program")
		#case for loading inputs
		if LOAD_SUCCESSFUL and useLoadedSettings:
			self.numMonomers = numMonomers
		else: 
			try:
				assert(self.monomerCountTkVar.get() > 0)
				global PENULTIMATE
				PENULTIMATE = self.penultimateTkVar.get()
			except:
				errorMessage("Please input valid parameters!", 220)
				return
		global COLORARRAY
		COLORARRAY = [COLOR1, COLOR2, COLOR3, COLOR4, COLOR5, COLOR6, COLOR7, COLOR8]
		#print("ColorArray: ", COLORARRAY)
		#Destroys or edits current widgets
		self.initFrame.destroy()
		self.loadButton.destroy()
		self.rowFrame1.destroy()
		self.rowFrame2.destroy()
		self.penultimateFrame.destroy()
		#Commands
		# Quit command: quits window
		def _quit():
			root.quit()
			root.destroy()
		# Back Command: goes back to numMonomers Entry
		def back(self):
			debug = False
			if debug:
				root.quit()
				root.destroy()
				return
			#Destroys all neccesary widgets
			self.destroyWidgets()
			#Re-Initiates
			self.initScreen()
			self.createInputWidgets()

		#Frame for totalMonomers label and Entry
		self.totalMonomersFrame = ttk.Frame(master = self.columnFrame)
		self.totalMonomersFrame.pack(side = Tk.TOP)
		#Label for totalMonomers Entry
		self.totalMonomersLabel = ttk.Label(master = self.totalMonomersFrame, text = "Monomer Pool Size:")
		self.totalMonomersLabel.pack(side = Tk.LEFT, pady = 3)
		#Entry for totalMonomers
		self.totalMonomersEntry = ttk.Entry(master = self.totalMonomersFrame, width = 7)
		self.totalMonomersEntry.pack(side = Tk.LEFT, padx = 3, pady = 3)
		#Setting totalMonomers to 1000
		self.totalMonomersTkVar = Tk.IntVar()
		self.totalMonomersEntry["textvariable"] = self.totalMonomersTkVar
		self.totalMonomersTkVar.set(TOTAL_STARTING_MONOMERS)
		#Frame for raftRatio label and Entry
		self.raftRatioFrame = ttk.Frame(master = self.columnFrame)
		self.raftRatioFrame.pack(side = Tk.TOP)
		#Label for raftRatio Entry
		self.raftRatioLabel = ttk.Label(master = self.raftRatioFrame, text = "Number Average DP:   ")
		self.raftRatioLabel.pack(side = Tk.LEFT, pady = 3)
		#Entry for raftRatio
		self.raftRatioEntry = ttk.Entry(master = self.raftRatioFrame, width = 5)
		self.raftRatioEntry.pack(side = Tk.LEFT, padx = 3, pady = 3)
		self.raftRatioTkVar = Tk.StringVar()
		self.raftRatioEntry["textvariable"] = self.raftRatioTkVar
		self.raftRatioTkVar.set(DP)
		#Frame for conversion
		self.conversionFrame = ttk.Frame(master = self.columnFrame)
		self.conversionFrame.pack(side = Tk.TOP, pady = 3)
		self.conversionLabel = ttk.Label(master = self.conversionFrame, text = "Percent Conversion:")
		self.conversionLabel.pack(side = Tk.LEFT)
		self.conversionTkVar = Tk.DoubleVar()
		self.conversionEntry = ttk.Entry(master	= self.conversionFrame, width = 4, textvariable = self.conversionTkVar)
		self.conversionTkVar.set(int(CONVERSION))
		self.conversionEntry.pack(side = Tk.LEFT)
		#Frame for numPolyToShow SpinBox
		self.numPolyToShowFrame = ttk.Frame(master = self.columnFrame)
		self.numPolyToShowFrame.pack(side = Tk.TOP, pady = 3)
		#Label for numPolyToShow spinbox
		self.numPolyToShowLabel = ttk.Label(master = self.numPolyToShowFrame, text = "Polymers to Show:")
		self.numPolyToShowLabel.pack(side = Tk.LEFT, padx = 0, pady = 0)
		#numPolyToShow spinbox
		self.numPolyToShowBox = Tk.Spinbox(master = self.numPolyToShowFrame, from_ = 1, to = 12, width = 2)
		self.numPolyToShowBox.pack(side = Tk.LEFT, padx = 5, pady = 0)
		#Setting numPolyToShow to 6
		self.numPolyToShow = Tk.IntVar()
		self.numPolyToShowBox["textvariable"] = self.numPolyToShow
		self.numPolyToShow.set(NUM_POLYMERS_SHOW)
		#Frame for Back, Quit, and Simulate buttons
		self.backSimFrame = ttk.Frame(master = self.columnFrame)
		self.backSimFrame.pack(side = Tk.TOP)
		#A simulate button to simulate polymer formation
		self.simulateButton = ttk.Button(master = self.backSimFrame, text = "Simulate", width = 8,
		 command = self.simulate2)
		self.simulateButton.pack(side = Tk.LEFT, padx = 2, pady = 4)
		#An Update Button Widget
		updateButton = ttk.Button(master = self.backSimFrame, text = "Update",
		 command = lambda:self.plotCompositions(True), width = 7)
		updateButton.pack(side = Tk.LEFT, padx = 2)
		#A save state button
		self.saveButton = ttk.Button(master = self.backSimFrame, text = "Save", width = 6,
		 command = self.saveState)
		self.saveButton.pack(side = Tk.LEFT, padx = 2, pady = 4)
		self.lastRowFrame = ttk.Frame(master = self.columnFrame)
		self.lastRowFrame.pack(side = Tk.TOP)
		#A back button to enter a diff number of monomers
		self.backButton = ttk.Button(master = self.lastRowFrame, text = "Back", width = 6,
		 command = lambda:back(self))
		self.backButton.pack(side = Tk.LEFT, padx = 2, pady = 4)
		#Export button
		self.exportButton = ttk.Button(master = self.lastRowFrame, text = "Export", width = 7, 
			command	= self.export)
		self.exportButton.pack(side = Tk.LEFT, padx = 2)	
		#Options button
		self.optionsButton = ttk.Button(master = self.lastRowFrame, text = "Options", width = 7, 
			command	= self.displayOptions)
		self.optionsButton.pack(side = Tk.LEFT, padx = 2)
		#seperator for column
		self.col1Sep = ttk.Separator(master = self.inputFrame, orient = Tk.VERTICAL)
		self.col1Sep.pack(side = Tk.LEFT, expand = True, fill = Tk.BOTH, pady = 1)
		#Frame for graph options
		self.column2Frame = ttk.Frame(master = self.inputFrame)
		self.column2Frame.pack(side = Tk.LEFT, padx = 8)
		#tkvar for graphType1
		self.graphType1TkVar = Tk.StringVar()
		#tkVar for graphType2
		self.graphType2TkVar = Tk.StringVar()
		#frame for graphType1
		self.graphFrame1 = ttk.Frame(master = self.column2Frame)
		self.graphFrame1.pack(side = Tk.TOP, pady = 3)
		#frame for graphType2
		self.graphFrame2 = ttk.Frame(master = self.column2Frame)
		self.graphFrame2.pack(side = Tk.TOP, pady = 3)
		#Label for graphType1 spinbox
		self.graphType1Label = ttk.Label(master = self.graphFrame1, text = "Graph 1 Type:")
		self.graphType1Label.pack(side = Tk.LEFT)
		#Label for graphType2 spinbox
		self.graphType2Label = ttk.Label(master = self.graphFrame2, text = "Graph 2 Type:")
		self.graphType2Label.pack(side = Tk.LEFT)
		#combobox
		self.graphType1ComboBox = ttk.Combobox(master = self.graphFrame1, values = ("Monomer Occurrences", "Monomer Usage", 
			"Run Length", "DP Distribution", "Hydrophobic Run Length", "Hydrophilic Run Length", "None"), 
		textvariable = self.graphType1TkVar, state = "readonly", width = 21)
		self.graphType1ComboBox.pack(side = Tk.LEFT)

		#Spinbox for graphType1Spinbox
		#self.graphType1Spinbox = Tk.Spinbox(master = self.graphFrame1, values = ("Monomer Occurrences", "Percentage Monomer", 
		#	"Monomer Separation", "None"), width = 20, textvariable = self.graphType1TkVar, state = "readonly")
		#self.graphType1Spinbox.pack(side = Tk.LEFT)
		#setting default variable for graphType1
		self.graphType1TkVar.set(GRAPH1_TYPE)
		#combobox
		self.graphType2ComboBox = ttk.Combobox(master = self.graphFrame2, values = ("Monomer Occurrences", "Monomer Usage", 
			"Run Length", "DP Distribution", "Hydrophobic Run Length", "Hydrophilic Run Length", "None"),
			 textvariable = self.graphType2TkVar, state = "readonly", width = 21)
		self.graphType2ComboBox.pack(side = Tk.LEFT)
		#Frame for histogramLimit
		self.histogramLimitFrame = ttk.Frame(master = self.column2Frame)
		self.histogramLimitFrame.pack(side = Tk.TOP)
		#Label for histogramLimit DEPRECATED
		self.histogramLimitLabel = ttk.Label(master = self.histogramLimitFrame, text = "Percentage to Analyze for Histogram:")
		#self.histogramLimitLabel.pack(side = Tk.LEFT)
		#entry for histogramLimit
		self.histogramLimitEntry = ttk.Entry(master = self.histogramLimitFrame, width = 4)
		#self.histogramLimitEntry.pack(side = Tk.LEFT)
		#tkvar for histogramLimit
		self.histogramLimitTkVar = Tk.StringVar()
		self.histogramLimitEntry["textvariable"] = self.histogramLimitTkVar
		self.histogramLimitTkVar.set(HISTOGRAM_LIMIT)
		#Spinbox for graphType2
		#self.graphType2Spinbox = Tk.Spinbox(master = self.graphFrame2, values = ("Monomer Occurrences", "Percentage Monomer", 
		#"Monomer Separation", "None"), width = 20, textvariable = self.graphType2TkVar, state = "readonly")
		#self.graphType2Spinbox.pack(side = Tk.LEFT)
		#setting default variable for graphType2
		self.graphType2TkVar.set(GRAPH2_TYPE)
		#Frame for histogramMonomer1
		self.histogramMonomer1Frame = ttk.Frame(master = self.column2Frame)
		self.histogramMonomer1Frame.pack(side = Tk.TOP)
		#Label for histogramMonomer1
		self.histogramMonomer1Label = ttk.Label(master = self.histogramMonomer1Frame, text = "Histogram 1 Monomer:")
		self.histogramMonomer1Label.pack(side = Tk.LEFT)
		#tkvar for histogramMonomer1
		self.histogramMonomer1TkVar = Tk.IntVar()
		#spinbox for histogramMonomer1
		self.histogramMonomer1Spinbox  = Tk.Spinbox(master = self.histogramMonomer1Frame, from_ = 1, to = self.numMonomers, width = 2, state = "readonly")
		self.histogramMonomer1Spinbox.pack(side = Tk.LEFT)
		#setting spinbox to default value
		self.histogramMonomer1Spinbox["textvariable"] = self.histogramMonomer1TkVar
		self.histogramMonomer1TkVar.set(HISTOGRAM1_MONOMER)
		#Frame for histogramMonomer2
		self.histogramMonomer2Frame = ttk.Frame(master = self.column2Frame)
		self.histogramMonomer2Frame.pack(side = Tk.TOP)
		#Label for histogramMonomer2
		self.histogramMonomer2Label = ttk.Label(master = self.histogramMonomer2Frame, text = "Histogram 2 Monomer:")
		self.histogramMonomer2Label.pack(side = Tk.LEFT)
		#tkvar for histogramMonomer2
		self.histogramMonomer2TkVar = Tk.IntVar()
		#spinbox for histogramMonomer2
		self.histogramMonomer2Spinbox  = Tk.Spinbox(master = self.histogramMonomer2Frame, from_ = 1, to = self.numMonomers, width = 2, state = "readonly")
		self.histogramMonomer2Spinbox.pack(side = Tk.LEFT)
		#setting spinbox to default value
		self.histogramMonomer2Spinbox["textvariable"] = self.histogramMonomer2TkVar
		self.histogramMonomer2TkVar.set(HISTOGRAM2_MONOMER)
		#seperator for column2
		self.col2Sep = ttk.Separator(master = self.inputFrame, orient = Tk.VERTICAL)
		self.col2Sep.pack(side = Tk.LEFT, expand = True, fill = Tk.BOTH, padx = 1, pady = 1)
		self.createIterativeInputs(False)
	def key(self):
		self.simulate2()
	
	"-------------------------------------------------------------------------------------------------------------------------------"
	"***CREATE ITERATIVE INPUTS: MONOMER RATIOS AND REACTIVITY RATIOS***"
	"-------------------------------------------------------------------------------------------------------------------------------"
	
	def createIterativeInputs(self, alias):
		root.bind("<Return>", lambda e: self.key())
		self.hphobList = []
		if HYDROPHOBICITY == 0:
			phobicity = "Hydrophobic"
		elif HYDROPHOBICITY == 1:
			phobicity = "Hydrophilic"
		else:
			phobicity = "None"
		for count in range(1,self.numMonomers + 1):
			self.hphobList.append(phobicity)
		#Frame for Monomer Amounts
		self.amountFrame = ttk.Frame(master = self.inputFrame) 
		self.amountFrame.pack(side = Tk.LEFT, padx = 0)
		#Frame for Monomer Coefficients
		self.coefficientFrame = ttk.Frame(master = self.inputFrame)
		self.coefficientFrame.pack(side = Tk.LEFT, padx = 0)

		# A list of ttk.Entry objects for Monomer ratios
		self.startingRatiosTkList = [] 
		# A 2D list of ttk.Entry objects for Coefficicients
		self.coefficientList = [] 
		#a list of the ttk.Entry objects for monomer input
		self.monomerAmountTkVarArray = []
		#a list of ttk.Label Objects for ratios
		self.ratiosLabelList = []
		#a list of ttk.Label Objects for coefficients
		self.coeffLabelList = []
		#variable for keeping track of monomer ratio loop
		createCount = 0;
		#While loop creating number of neccesary amount Entry boxes
		while createCount < self.numMonomers:
			#Label for inputAmount
			monomerAmountFrame = ttk.Frame(master = self.amountFrame)
			monomerAmountFrame.pack(side = Tk.TOP, padx = 0, pady = 3)
			inputAmountLabel = ttk.Label(master = monomerAmountFrame, text = "     Monomer " 
				+ str(createCount + 1) + " Percentage: ")
			self.ratiosLabelList.append(inputAmountLabel)
			inputAmountLabel.pack(side = Tk.LEFT)
			#Entry for inputAmount
			amount = Tk.IntVar()
			inputAmount = ttk.Entry(master = monomerAmountFrame, width = 5, textvariable = amount, text = 1)
			inputAmount.pack(side = Tk.LEFT, padx = 2)
			#Setting Default Value to 1
			inputAmount["textvariable"] = amount
			if LOAD_SUCCESSFUL and useLoadedSettings:
				amount.set(RATIO_ARRAY[createCount])
			else:
				amount.set(1)
			self.monomerAmountTkVarArray.append(amount)
			#Add ttk.Entry object to startingAmountList
			self.startingRatiosTkList.append(inputAmount)
			createCount += 1
		setDefaultCount = 0
		"""while setDefaultCount < self.numMonomers:
			self.startingRatiosTkList[setDefaultCount].configure(text = 20)
			setDefaultCount += 1"""
		#Debugging purposes
		#print("startingAmountList: ", self.startingRatiosTkList) 
		if not PENULTIMATE:
			#case for 2 monomer system - only need 2 reactivity ratios
			if self.numMonomers == 2:
				self.coeffTkVarArray = []
				monomerIDcount = 0
				while monomerIDcount < 2:
					singleMonoCoeffList = []
					self.coefficientList.append(singleMonoCoeffList)
					self.coeffLabelList.append([])
					count = 0
					while count < 2:			
						#Label for inputAmount
						coeffValFrame = ttk.Frame(master = self.coefficientFrame)
						if count == monomerIDcount:
							coeffValFrame.pack(side = Tk.TOP, padx = 2, pady = 3)
						inputCoeffLabel = ttk.Label(master = coeffValFrame, text = "Reactivity Ratio %i:" %(monomerIDcount + 1) )
						self.coeffLabelList[monomerIDcount].append(inputCoeffLabel)
						if count == monomerIDcount:
							inputCoeffLabel.pack(side = Tk.LEFT)
						#Entry for inputAmount
						inputCoeff = ttk.Entry(master = coeffValFrame, width = 4)
						if count == monomerIDcount:
							inputCoeff.pack(side = Tk.LEFT, padx = 5)
						#Setting Default Coefficient to 1
						coeff = Tk.IntVar()
						inputCoeff["textvariable"] = coeff
						if LOAD_SUCCESSFUL and useLoadedSettings:
							coeff.set(COEFF_ARRAY[monomerIDcount][count])
						else:
							coeff.set(1)
						#Ensures that hetegenous coefficients are always 1
						#if count == 2:
						#	coeff.set(1)
						#IMPORTANT: saves input so won't be garbage collected. DO NOT DELETE THIS LINE< EVEN IF IT SEEMS USELESS
						self.coeffTkVarArray.append(coeff)
						#Add a ttk.Entry object to singleMonoCoeffList
						singleMonoCoeffList.append(inputCoeff)
						count += 1
					monomerIDcount += 1
			else:
				#IMPORTANT: list so that inputs can be udated and not trash collected
				self.coeffTkVarArray = []
				createCount2 = 0
				#While loop creating number of neccesary coefficient Entry boxes
				for firstMonomer in range(1, self.numMonomers + 1):
					combinations = 0
					#Appends to coefficient list a list containing coefficients for the polymer index
					singleMonoCoeffList = []
					self.coefficientList.append(singleMonoCoeffList)
					self.coeffLabelList.append([])
					#Frame for Coefficients for Single Monomer
					singleCoeffFrame = ttk.Frame(master = self.coefficientFrame)
					singleCoeffFrame.pack(side = Tk.LEFT, fill = Tk.X, expand = 1)
					coeffID = 0
					for secondMonomer in range(1, self.numMonomers + 1):			
						#Label for inputAmount
						if secondMonomer != firstMonomer:
							coeffValFrame = ttk.Frame(master = singleCoeffFrame)
							coeffValFrame.pack(side = Tk.TOP, padx = 2, pady = 3)
							inputCoeffLabel = ttk.Label(master = coeffValFrame, text = " rr" + str(firstMonomer) 
								+ str(secondMonomer) + ": ")
							self.coeffLabelList[createCount2].append(inputCoeffLabel)
							inputCoeffLabel.pack(side = Tk.LEFT)
							#Entry for inputAmount
							inputCoeff = ttk.Entry(master = coeffValFrame, width = 4)
							inputCoeff.pack(side = Tk.LEFT, padx = 5)
							#Setting Default Coefficient to 1
							coeff = Tk.IntVar()
							inputCoeff["textvariable"] = coeff
							if LOAD_SUCCESSFUL and useLoadedSettings:
								print("COEFF_ARRAY: ", COEFF_ARRAY)
								coeff.set(COEFF_ARRAY[firstMonomer - 1][coeffID])
							else:
								coeff.set(1)
							coeffID += 1
							#IMPORTANT: saves input so won't be garbage collected. DO NOT DELETE THIS LINE< EVEN IF IT SEEMS USELESS
							self.coeffTkVarArray.append(coeff)
							#Add a ttk.Entry object to singleMonoCoeffList
							singleMonoCoeffList.append(inputCoeff)
		#if PENULTIMATE is true, create input entries for penultimate model
		#format: penultimate-ultimate-next , so 1-2-3 is 1 (penultimate), 2 (ultimate), 3 (next)
		if PENULTIMATE:
			self.coeffTkVarArray = []
			#for loop creating neccesary number of penultimate coefficient entry boxes
			for nextMonomer in range(0, self.numMonomers):
				nextMonomerList = []
				self.coefficientList.append(nextMonomerList)
				#Frame for Coefficients for Single Monomer
				singleCoeffFrame = ttk.Frame(master = self.coefficientFrame)
				singleCoeffFrame.pack(side = Tk.LEFT, fill = Tk.X, expand = 1)
				for ultimate in range(0, self.numMonomers):
					CoeffList = []
					nextMonomerList.append(CoeffList)
					for penultimate in range(0, self.numMonomers):
						#Label for inputAmount
						coeffValFrame = ttk.Frame(master = singleCoeffFrame)
						coeffValFrame.pack(side = Tk.TOP, padx = 5, pady = 3)
						if alias:
							inputCoeffLabel = ttk.Label(master = coeffValFrame, text = self.aliasList[penultimate] + "-" + self.aliasList[ultimate]
						 + "-" + self.aliasList[nextMonomer] + " Constant:" )
						inputCoeffLabel = ttk.Label(master = coeffValFrame, text = str(penultimate + 1) + "-" + str(ultimate + 1)
						 + "-" + str(nextMonomer + 1) + " Constant:" )
						inputCoeffLabel.pack(side = Tk.LEFT)
						#Entry for inputAmount
						inputCoeff = ttk.Entry(master = coeffValFrame, width = 4)
						inputCoeff.pack(side = Tk.LEFT, padx = 5)
						#Setting Default Coefficient to 1
						coeff = Tk.IntVar()
						inputCoeff["textvariable"] = coeff
						if LOAD_SUCCESSFUL and useLoadedSettings:
							coeff.set(COEFF_ARRAY[nextMonomer][ultimate][penultimate])
						else:
							coeff.set(1)
						self.coeffTkVarArray.append(coeff)
						#Add a ttk.Entry object to singleMonoCoeffList
						CoeffList.append(inputCoeff)

	"-----------------------------------------------------------------------------------------------------------------------------------------------------"
	"***MAIN CODE FOR SIMULATING POLYMERS***"
	"-----------------------------------------------------------------------------------------------------------------------------------------------------"

	#Simulates polymer reaction based on input values
	def simulate(self):
		root.wm_state('zoomed')
		#print("Simulating!")
		#a lock
		if self.simulateLocked == True:
			return

		"***Read user inputs and checks to see if they are valid***"
		try:
			self.totalMonomers = int(self.totalMonomersTkVar.get())
			monomerAmounts = self.getMonomerAmounts()
			if not PENULTIMATE:
				coeffList = self.getCoefficients()
				singleCoeffList = self.getReactivity(coeffList, self.numMonomers)
			else:
				singleCoeffList = self.getPenultimateCoeff()
			self.raftRatio = float(self.raftRatioTkVar.get())
			self.histogramLimit = float(self.histogramLimitTkVar.get())
			self.conversion = float(self.conversionTkVar.get())
			assert(self.histogramLimit <= 1)
			assert(self.conversion > 0)
			assert(self.conversion <= 100)
			assert(self.histogramLimit > 0)
			assert(self.totalMonomers > 0)
			assert(NUM_SIMULATIONS > 0)
			assert(self.raftRatio > 0)
			assert(self.totalMonomers * NUM_SIMULATIONS <= MONOMER_CAP)
			self.graph1Type = self.graphType1TkVar.get()
			self.graph2Type = self.graphType2TkVar.get()
			if self.graph1Type == "Hydrophobic Blocks" or self.graph1Type == "Hydrophilic Blocks" or self.graph2Type == "Hydrophobic Blocks" or self.graph2Type == "Hydrophilic Blocks":
				for phobicity in self.hphobList:
					if phobicity == "None":
						raise phobicityNotSpecified("hi")
		except ValueError:
			errorMessage("Please input valid parameters!", 220)
			return
		except notInEuropeError:
			errorMessage("You are not in Europe!", 220)
			return
		except AssertionError:
			errorMessage("Please input valid parameters!", 220)
			return
		except phobicityNotSpecified:
			errorMessage("Please specficy hydrophobicities in the Options Tab.", 330)
			return
		self.numSimulations = NUM_SIMULATIONS
		self.fullPolymerLength = int(self.raftRatio)

		"***Calculate the polymer length using raft ratio and percent conversion***"
		self.polymerLength = int(self.raftRatio * self.conversion / 100)
		if self.polymerLength == 0:
			self.polymerLength = 1

		"***Caluculate number of polymer chains by dividing total number of monomers by monomers per full polymer***"
		self.numPolymers = int(self.totalMonomers / self.fullPolymerLength)
		#Asserting valid inputs for RAFT ration and totalMonomers
		try:
			assert self.numPolymers > 0
		except AssertionError:
			errorMessage("RAFT Ratio too small!", 220)
			return
		#print("numMonomers: ", self.numMonomers)


		
		#Progress bar widget setup
		self.simulateLocked = True
		simulationProgressTkVar = Tk.IntVar()
		self.simulationProgressBar = ttk.Progressbar(master = self.column2Frame, mode = "determinate", 
			variable = simulationProgressTkVar, maximum = self.polymerLength * NUM_SIMULATIONS, length = 200)
		self.simulationProgressBar.pack(side = Tk.TOP, padx = 5, pady = 2)
		self.simulationProgressBar.update()
		simulationProgressTkVar.set(0)
		currProgress = 0
		self.simulationProgressBar.start()

		#number of monomers in each polymer assuming reaction goes to completion, based on raftRatio and itotalMonomers
		#print(self.totalMonomers)
		#print(self.raftRatio)

		if self.initialSetTkVar.get() == "Weighted":
			global SETINITIAL
			SETINITIAL = 0
		else:
			wordarray = self.initialSetTkVar.get().split()
			global SETINITIAL
			SETINITIAL = int(wordarray[1])

		"***initiate an empty array, which will hold all of the polymers to be simulate***"
		self.polymerArray = []
		#keeps track of number of simulations
		simCounter = 1
		self.originalMonomerAmounts = list(monomerAmounts)
		while simCounter <= self.numSimulations:
			#a local array of polymers
			localPolymerArray = []
			#sets monomerAmounts to orignalMonomerAmounts
			monomerAmounts = list(self.originalMonomerAmounts)
			#print("originalMonomerAmounts: ", originalMonomerAmounts)
			#variable keeping track of current number of polymers
			currNumPolymers = 0
			#print(self.numMonomers)

			"-----------------------------------------------------------------------------------------------------------------------"
			"***INITIATION STEP***"
			"------------------------------------------------------------------------------------------------------------------------"

			"In this step, all polymers will be initiated with one starting monomer. The algorithm continually intiates"
			"a chain until the chain count reaches self.numPolymers."

			while currNumPolymers < self.numPolymers:  

				"In the case where the user fixes all chains to start with a certain monomer, we simply initiate all chains"
				"with the monomer of choice."
				
				#case for setting first monomer
				if SETINITIAL:
					localPolymerArray.append([SETINITIAL])
					monomerAmounts[SETINITIAL - 1] -= 1
					currNumPolymers += 1
					continue

				"Otherwise, we will choose the initial monomer randomly based on the the mole fractions 'f' of each monomer"
				"in the feed solution. For 2- and 3- monomer systems, the instantaneous form of the Mayo Lewis Equation is "
				"used to determine the probabilities of each monomer initiating the chain."

				"Initiate a variable 'choices' that keeps track of the weight probabilty assigned to each monomer."
				#Example:

				#>>>print(choices)
				#[[1, 1.5], [2, 0.5]]

				#This means that monomer 1 has a weight of 1.5 assigned to it and monomer 2 has a weight of 0.5 assigned to it.
				#Thus, monomer 1 will initiate 3x more often than monomer 2 will.

				choices = []

				"Case for 2-monomer system: use the Mayo Lewis Equation"

				if self.numMonomers == 2 and not PENULTIMATE:
					f1 = monomerAmounts[0]
					f2 = monomerAmounts[1]
					r1 = singleCoeffList[0][0]
					r2 = singleCoeffList[1][1]
					#print("r1: ", r1)
					#print("r2: ", r2)
					weight = (r1*f1**2 + f1*f2) / (r1*f1**2 + 2*f1*f2 + r2*f2**2)
					#print("weight: ", weight)
					choices.append([1, weight])
					choices.append([2, 1 - weight])

				

				elif self.numMonomers == 3:
					"Case for 3-monomer system: use an altered Mayo Lewis Equation"
					m1 = monomerAmounts[0]
					m2 = monomerAmounts[1]
					m3 = monomerAmounts[2]
					F = m1 + m2 + m3
					f1 = m1/F
					f2 = m2/F
					f3 = m3/F
					r11 = singleCoeffList[0][0]
					r12 = singleCoeffList[0][1]
					r13 = singleCoeffList[0][2]
					r21 = singleCoeffList[1][0]
					r22 = singleCoeffList[1][1]
					r23 = singleCoeffList[1][2]
					r31 = singleCoeffList[2][0]
					r32 = singleCoeffList[2][1]
					r33 = singleCoeffList[2][2]
					R1 = r11 + r12 + r13
					R2 = r21 + r22 + r23
					R3 = r31 + r32 + r33
					a = f1*r11*f1/(r11*f1+r12*f2+r13*f3) + f2*r21*f1/(r21*f1+r22*f2+r23*f3) + f3*r31*f1/(r31*f1+r32*f2+r33*f3)
					b = f1*r12*f2/(r11*f1+r12*f2+r13*f3) + f2*r22*f2/(r21*f1+r22*f2+r23*f3) + f3*r32*f2/(r31*f1+r32*f2+r33*f3)
					c = 1 - a - b
					#print("startingRatioList: ", [a, b, c])
					choices.append([1,a])
					choices.append([2 ,b])
					choices.append([3,c])
				else:
					"Case for any monomer system > 3: initial solely based on feed mole ratios 'f'"
					#A variable keeping track of current monomer
					monomerID = 1

					"Cycle through each monomer, finding its feed amount and using that value as the weight."

					while monomerID <= self.numMonomers:
						#weight chance of monomer initation: (amount of starting monomer)
						weight = monomerAmounts[monomerID - 1]
						#Adds a two element list to choices containing monomer and weight
						choices.append([monomerID, weight])
						monomerID += 1

				"Randomly choose a monomer to initiate the polymer chain by using a weighted random selector."
				"Monomer with higher relative weights will be chosen more often. The 'weighted_choices' function"
				"takes in the 'choices' variable which has relevant weights for each monomer and runs a weighted"
				"random selection on it."

				startingMonomer = weighted_choice(choices)
				#Starts a new polymer with startingMonomer, represented by an array, 
				#and adds that array to polymerArray
				"A polymer is represented as a python list/ array. Here, we inititate an instance of a polymer as a list, "
				"and append that polymer to a superlist which contains a list of all polymers"

				localPolymerArray.append([startingMonomer])

				"Remove the monomer that was used to intitate the polymer in order to accurately update the monomer pool ratios."
				#Uses up one monomer of the startingMonomer if MAINTAIN is false, else maintains composition
				if not MAINTAIN:
					monomerAmounts[startingMonomer - 1] -= 1
				#increases number of polymers by 1
				currNumPolymers += 1
			#debugging starting monomers
			#print("polymer array", polymerArray)
			#variable to keep track of polymer length
			currPolymerLength = 1
			"""Grows each polymer at the same time until they all reach desired polymer size"""
			#print("self.polymerLength: ", self.polymerLength)

			"case for penultimate model: choosing the second monomer in chain, based off of "
			"HETEROGENOUS constants (works only for 2 monomer system)"

			if PENULTIMATE:
				#iterating through each polymer in batch
				for polymer in localPolymerArray:
					choices = []
					#calculates weight chance for each monomer
					for monomerID in range(1, self.numMonomers + 1):
						#retrieving coefficient based on previous and currrent monomer, assumes hetergoenous penultimate monomer
						coeff = singleCoeffList[monomerID - 1][polymer[-1] - 1][1]
						#print("coeff: ", coeff)
						# weight chance calulations for monomer attaching: coefficient*(amount of monomer remaining)
						chance = monomerAmounts[monomerID - 1] * coeff
						#adds a two element list to choices containing monomer and weight
						choices.append([monomerID, chance])
					#Using weighted_choice, selects next monomer
					try:
						#print(choices)
						nextMonomer = weighted_choice(choices)	
					#If all weights are zero due to coefficients, then sort by relative amounts of monomer instead
					except AssertionError:
						monomerID = 1
						choices = []
						while monomerID <= self.numMonomers:
							choices.append([monomerID, monomerAmounts[monomerID - 1]])
							monomerID += 1
						nextMonomer = weighted_choice(choices)
					#Reduces number of nextMonomer by 1, since it is being used up in reaction, unless MAINTAIN is true
					if not MAINTAIN:
						monomerAmounts[nextMonomer - 1] -= 1
					#print("monomerAmounts: ", monomerAmounts)
					#Attaches next monomer to polymer chain
					polymer.append(nextMonomer)

			"-------------------------------------------------------------------------------------------------------------------"
			"***PROPAGATION STEP***"
			"-------------------------------------------------------------------------------------------------------------------"

			"Case for penultimate step (description will not be as in depth). For normal Mayo-Lewis case, scroll down."

			if PENULTIMATE:
				for currPolymer in range(2, self.polymerLength):	
					for polymer in localPolymerArray:
						choices = []
						#calculates weight chance for each monomer
						for monomerID in range(1, self.numMonomers + 1):
							#retrieving coefficient based on previous and currrent monomer, assumes hetergoenous penultimate monomer
							#assigning variables for clarity
							nextMonomer = monomerID - 1
							ultimateMonomer = polymer[-1] - 1
							penultimateMonomer = polymer[-2] - 1
							coeff = singleCoeffList[nextMonomer][ultimateMonomer][penultimateMonomer]
							# weight chance calulations for monomer attaching: coefficient*(amount of monomer remaining)
							chance = monomerAmounts[monomerID - 1] * coeff
							#adds a two element list to choices containing monomer and weight
							choices.append([monomerID, chance])
						#Using weighted_choice, selects next monomer
						try:
							nextMonomer = weighted_choice(choices)
						#If all weights are zero due to coefficients, then sort by relative amounts of monomer instead
						except AssertionError:
							monomerID = 1
							choices = []
							while monomerID <= self.numMonomers:
								choices.append([monomerID, monomerAmounts[monomerID - 1]])
								monomerID += 1
							nextMonomer = weighted_choice(choices)
						#Reduces number of nextMonomer by 1, since it is being used up in reaction, unless MAINTAIN is true
						if not MAINTAIN:
							monomerAmounts[nextMonomer - 1] -= 1
						#print("monomerAmounts: ", monomerAmounts)
						#Attaches next monomer to polymer chain
						polymer.append(nextMonomer)
					#increase progress
					currProgress += 1
					simulationProgressTkVar.set(currProgress)
					self.simulationProgressBar.update()

			
			else:
				"***Propogation for standard Mayo-Lewis Case***"

				"The 'while' loop cycles through each round of propogation until the target degree of polymerization is met. "
				while currPolymerLength < self.polymerLength:
					"The 'for' loop cycles between all polymers (instantiated in the previous initiation step), and for each"
					"polymer, it appends one monomer to the growing chain, simulating an ideal homogenous polymer chain growth."
					for polymer in localPolymerArray:
						#A variable keeping track of current monomer
						monomerID = 1
						choices = []
						"For each polymer, iterate through all possible monomers which can be added. For each monomer, calculate"
						"the weight chance of the monomer to be added to the chain , defined as the product of the relevant rate "
						"constant 'k' times the number of monomers left unreacted 'f'"
						while monomerID <= self.numMonomers:
							#Retrieveing coefficient based on previous and current monomer
							#get the last monomer on the growing chain
							terminatingMonomer = polymer[-1]
							#retrieve the relevant rate constant k
							k = singleCoeffList[terminatingMonomer - 1][monomerID - 1]
							# weight chance calulations for monomer attaching: coefficient*(amount of monomer remaining)
							chance = monomerAmounts[monomerID - 1] * k
							#print(monomerID, " amount of monomer remaining: ", monomerAmounts[monomerID - 1]);
							#Adds a two element list to choices containing monomer and weight
							choices.append([monomerID, chance])
							monomerID += 1
						#print ("currPolymerLength: ", currPolymerLength)
						#print("choices2: ", choices)
						#Using weighted_choice, selects next monomer
						try:
							nextMonomer = weighted_choice(choices)
						#If all weights are zero due to coefficients, then sort by relative amounts of monomer instead
						except AssertionError:
							monomerID = 1
							choices = []
							while monomerID <= self.numMonomers:
								choices.append([monomerID, monomerAmounts[monomerID - 1]])
								monomerID += 1
							nextMonomer = weighted_choice(choices)
						#Reduces number of nextMonomer by 1, since it is being used up in reaction, unless MAINTAIN is true
						if not MAINTAIN:
							monomerAmounts[nextMonomer - 1] -= 1
						#print("monomerAmounts: ", monomerAmounts)
						#Attaches next monomer to polymer chain
						polymer.append(nextMonomer)
					#increase current polymer length by 1
					currPolymerLength += 1
					#increase progress
					currProgress += 1
					simulationProgressTkVar.set(currProgress)
					self.simulationProgressBar.update()
					#print("currProgress: ", currProgress)
			#print("simCounter: ", simCounter)
			#adds local polymer array to global polymer array
			self.polymerArray = self.polymerArray + localPolymerArray
			#increases simCounter by 1
			simCounter += 1


		"""Important Debug: Prints out array of polymers"""
		#print("Array of Polymers: ", polymerArray)

		#stops progress bar
		self.simulationProgressBar.stop()

		"--------------------------------------------------------------------------------------------------------------------"
		"***OUTPUT RAW SIMULATION RESULTS INTO A TEXT FILE"
		"--------------------------------------------------------------------------------------------------------------------"

		text_file = open("polymerArray.txt", "w")
		json.dump(self.polymerArray, text_file)
		text_file.close()

		"--------------------------------------------------------------------------------------------------------------------"
		"***CALCULATE AND DISPLAY MONOMER COMPOSITION***"
		"--------------------------------------------------------------------------------------------------------------------"

		self.compositionList = self.getComposition(self.polymerArray)
		if not self.compFrameExists:
			self.col3Sep = ttk.Separator(master = self.inputFrame, orient = Tk.VERTICAL)
			self.col3Sep.pack(side = Tk.LEFT, expand = True, fill = Tk.BOTH, pady = 1)
			self.compFrame = ttk.Frame(master = self.inputFrame)
			self.compFrame.pack(side = Tk.LEFT, padx = 0)
			self.compTkVarArray = []
			self.compLabelList = []
			for monomer in range(1, self.numMonomers + 1):
				dispCompFrame = ttk.Frame(master = self.compFrame)
				dispCompFrame.pack(side = Tk.TOP, pady = 3)
				if ALIAS:
					label = self.aliasList[monomer - 1]
				else:
					label = "Monomer " + str(monomer)
				fullLabel = label + " Composition: "
				compLabel = ttk.Label(master = dispCompFrame, text = fullLabel)
				self.compLabelList.append(compLabel)
				compLabel.pack(side = Tk.LEFT)
				compTkVar = Tk.DoubleVar()
				compTkVar.set(round(self.compositionList[monomer - 1], 2))
				self.compTkVarArray.append(compTkVar)
				compEntry = ttk.Entry(master = dispCompFrame, textvariable = compTkVar, width = 4)
				compEntry.pack(side = Tk.LEFT, padx = 5)
				self.compFrameExists = True
		else:
			monomerID = 1
			for item in self.compTkVarArray:
				item.set(round(self.compositionList[monomerID - 1], 2))
				monomerID += 1

		"***Visualize Polymers***"
		self.visualizePolymers(self.polymerArray)

		"***Plot Polymer Data***"
		self.plotCompositions(False)
		#destroy progress bar
		self.simulationProgressBar.destroy()
		center(root)
		self.simulateLocked = False
		#self.inputFrame.pack_forget()
	"-----------------------------------------------------------------------------------------------------------------------------------------------------"
	"***MAIN CODE FOR SIMULATING POLYMERS WITH VARYING LENGTHS***"
	"-----------------------------------------------------------------------------------------------------------------------------------------------------"

	#Simulates polymer reaction based on input values
	def simulate2(self):
		root.wm_state('zoomed')
		#print("Simulating!")
		#a lock
		if self.simulateLocked == True:
			return
		self.simulateLocked = True
		"***Read user inputs and checks to see if they are valid***"
		try:
			self.totalMonomers = int(self.totalMonomersTkVar.get())
			monomerAmounts = self.getMonomerAmounts()
			if not PENULTIMATE:
				coeffList = self.getCoefficients()
				singleCoeffList = self.getReactivity(coeffList, self.numMonomers)
			else:
				singleCoeffList = self.getPenultimateCoeff()
			self.raftRatio = float(self.raftRatioTkVar.get())
			self.histogramLimit = float(self.histogramLimitTkVar.get())
			self.conversion = float(self.conversionTkVar.get())
			assert(self.histogramLimit <= 1)
			assert(self.conversion > 0)
			assert(self.conversion <= 100)
			assert(self.histogramLimit > 0)
			assert(self.totalMonomers > 0)
			assert(NUM_SIMULATIONS > 0)
			assert(self.raftRatio > 0)
			assert(self.totalMonomers * NUM_SIMULATIONS <= MONOMER_CAP)
			self.graph1Type = self.graphType1TkVar.get()
			self.graph2Type = self.graphType2TkVar.get()
			if self.graph1Type == "Hydrophobic Blocks" or self.graph1Type == "Hydrophilic Blocks" or self.graph2Type == "Hydrophobic Blocks" or self.graph2Type == "Hydrophilic Blocks":
				for phobicity in self.hphobList:
					if phobicity == "None":
						raise phobicityNotSpecified("hi")
		except ValueError:
			errorMessage("Please input valid parameters!", 220)
			return
		except notInEuropeError:
			errorMessage("You are not in Europe!", 220)
			return
		except AssertionError:
			errorMessage("Please input valid parameters!", 220)
			return
		except phobicityNotSpecified:
			errorMessage("Please specficy hydrophobicities in the Options Tab.", 330)
			return
		self.numSimulations = NUM_SIMULATIONS
		self.fullPolymerLength = int(self.raftRatio)

		"***Calculate the polymer length using raft ratio and percent conversion***"
		self.polymerLength = int(self.raftRatio * self.conversion / 100)
		if self.polymerLength == 0:
			self.polymerLength = 1

		"***Caluculate number of polymer chains by dividing total number of monomers by monomers per full polymer***"
		self.numPolymers = int(self.totalMonomers / self.fullPolymerLength)
		#Asserting valid inputs for RAFT ration and totalMonomers
		try:
			assert self.numPolymers > 0
		except AssertionError:
			errorMessage("RAFT Ratio too small!", 220)
			return
		#print("numMonomers: ", self.numMonomers)

		if self.initialSetTkVar.get() == "Weighted":
			global SETINITIAL
			SETINITIAL = 0
		else:
			wordarray = self.initialSetTkVar.get().split()
			global SETINITIAL
			SETINITIAL = int(wordarray[1])

		"***initiate an empty array, which will hold all of the polymers to be simulated***"
		self.polymerArray = []
		#keeps track of number of simulations
		simCounter = 1
		self.originalMonomerAmounts = list(monomerAmounts)
		#sets monomerAmounts to orignalMonomerAmounts
		monomerAmounts = list(self.originalMonomerAmounts)
		totalOriginalMonomerAmounts = sum(list(self.originalMonomerAmounts))
		#print("monomerAmounts: ", monomerAmounts)
		#variable keeping track of current number of polymers
		currNumPolymers = 0
		#print(self.numMonomers)
		self.monomerUsageList = []
		self.monomerRemainingList = []
		for i in range(self.numMonomers):
			self.monomerUsageList.append([])
			self.monomerRemainingList.append([])
		self.monomerUsageSubList = [0]*self.numMonomers

		"-----------------------------------------------------------------------------------------------------------------------"
		"***INITIATION STEP***"
		"------------------------------------------------------------------------------------------------------------------------"

		"In this step, all polymers will be initiated with one starting monomer. The algorithm continually intiates"
		"a chain until the chain count reaches self.numPolymers."


		while currNumPolymers < self.numPolymers:  

			"In the case where the user fixes all chains to start with a certain monomer, we simply initiate all chains"
			"with the monomer of choice."
			
			#case for setting first monomer
			if SETINITIAL:
				self.polymerArray.append([SETINITIAL])
				monomerAmounts[SETINITIAL - 1] -= 1
				currNumPolymers += 1
				continue

			"Otherwise, we will choose the initial monomer randomly based on the the mole fractions 'f' of each monomer"
			"in the feed solution. For 2- and 3- monomer systems, the instantaneous form of the Mayo Lewis Equation is "
			"used to determine the probabilities of each monomer initiating the chain."

			"Initiate a variable 'choices' that keeps track of the weight probabilty assigned to each monomer."
			#Example:

			#>>>print(choices)
			#[[1, 1.5], [2, 0.5]]

			#This means that monomer 1 has a weight of 1.5 assigned to it and monomer 2 has a weight of 0.5 assigned to it.
			#Thus, monomer 1 will initiate 3x more often than monomer 2 will.

			choices = []

			"Case for 2-monomer system: use the Mayo Lewis Equation"

			if self.numMonomers == 2 and not PENULTIMATE:
				f1 = monomerAmounts[0]
				f2 = monomerAmounts[1]
				r1 = singleCoeffList[0][0]
				r2 = singleCoeffList[1][1]
				#print("r1: ", r1)
				#print("r2: ", r2)
				weight = (r1*f1**2 + f1*f2) / (r1*f1**2 + 2*f1*f2 + r2*f2**2)
				#print("weight: ", weight)
				choices.append([1, weight])
				choices.append([2, 1 - weight])

			

			elif self.numMonomers == 3:
				"Case for 3-monomer system: use an altered Mayo Lewis Equation"
				m1 = monomerAmounts[0]
				m2 = monomerAmounts[1]
				m3 = monomerAmounts[2]
				F = m1 + m2 + m3
				f1 = m1/F
				f2 = m2/F
				f3 = m3/F
				r11 = singleCoeffList[0][0]
				r12 = singleCoeffList[0][1]
				r13 = singleCoeffList[0][2]
				r21 = singleCoeffList[1][0]
				r22 = singleCoeffList[1][1]
				r23 = singleCoeffList[1][2]
				r31 = singleCoeffList[2][0]
				r32 = singleCoeffList[2][1]
				r33 = singleCoeffList[2][2]
				R1 = r11 + r12 + r13
				R2 = r21 + r22 + r23
				R3 = r31 + r32 + r33
				a = f1*r11*f1/(r11*f1+r12*f2+r13*f3) + f2*r21*f1/(r21*f1+r22*f2+r23*f3) + f3*r31*f1/(r31*f1+r32*f2+r33*f3)
				b = f1*r12*f2/(r11*f1+r12*f2+r13*f3) + f2*r22*f2/(r21*f1+r22*f2+r23*f3) + f3*r32*f2/(r31*f1+r32*f2+r33*f3)
				c = 1 - a - b
				#print("startingRatioList: ", [a, b, c])
				choices.append([1,a])
				choices.append([2 ,b])
				choices.append([3,c])
			else:
				"Case for any monomer system > 3: initial solely based on feed mole ratios 'f'"
				#A variable keeping track of current monomer
				monomerID = 1

				"Cycle through each monomer, finding its feed amount and using that value as the weight."

				while monomerID <= self.numMonomers:
					#weight chance of monomer initation: (amount of starting monomer)
					weight = monomerAmounts[monomerID - 1]
					#Adds a two element list to choices containing monomer and weight
					choices.append([monomerID, weight])
					monomerID += 1

			"Randomly choose a monomer to initiate the polymer chain by using a weighted random selector."
			"Monomer with higher relative weights will be chosen more often. The 'weighted_choices' function"
			"takes in the 'choices' variable which has relevant weights for each monomer and runs a weighted"
			"random selection on it."

			startingMonomer = weighted_choice(choices)
			#Starts a new polymer with startingMonomer, represented by an array, 
			#and adds that array to polymerArray
			"A polymer is represented as a python list/ array. Here, we inititate an instance of a polymer as a list, "
			"and append that polymer to a superlist which contains a list of all polymers"

			self.polymerArray.append([startingMonomer])

			"Remove the monomer that was used to intitate the polymer in order to accurately update the monomer pool ratios."
			#Uses up one monomer of the startingMonomer if MAINTAIN is false, else maintains composition
			if not MAINTAIN:
				monomerAmounts[startingMonomer - 1] -= 1
			#increases number of polymers by 1
			currNumPolymers += 1
			index = 0
			for amount in monomerAmounts:
				self.monomerRemainingList[index].append(amount/ self.originalMonomerAmounts[index])
				index += 1
			self.monomerUsageSubList[startingMonomer - 1] += 1
		self.monomerUsageSubList = [float(i)/sum(self.monomerUsageSubList) for i in self.monomerUsageSubList]
		for i in range(self.numMonomers):
			self.monomerUsageList[i].append(self.monomerUsageSubList[i])


			
		#debugging starting monomers
		#print("polymer array", polymerArray)
		#variable to keep track of polymer length
		currPolymerLength = 1
		"""Grows each polymer at the same time until they all reach desired polymer size"""
		#print("self.polymerLength: ", self.polymerLength)

		"case for penultimate model: choosing the second monomer in chain, based off of "
		"HETEROGENOUS constants (works only for 2 monomer system)"

		if PENULTIMATE:
			#iterating through each polymer in batch
			for polymer in self.polymerArray:
				choices = []
				#calculates weight chance for each monomer
				for monomerID in range(1, self.numMonomers + 1):
					#retrieving coefficient based on previous and currrent monomer, assumes hetergoenous penultimate monomer
					coeff = singleCoeffList[monomerID - 1][polymer[-1] - 1][1]
					#print("coeff: ", coeff)
					# weight chance calulations for monomer attaching: coefficient*(amount of monomer remaining)
					chance = monomerAmounts[monomerID - 1] * coeff
					#adds a two element list to choices containing monomer and weight
					choices.append([monomerID, chance])
				#Using weighted_choice, selects next monomer
				try:
					#print(choices)
					nextMonomer = weighted_choice(choices)	
				#If all weights are zero due to coefficients, then sort by relative amounts of monomer instead
				except AssertionError:
					monomerID = 1
					choices = []
					while monomerID <= self.numMonomers:
						choices.append([monomerID, monomerAmounts[monomerID - 1]])
						monomerID += 1
					nextMonomer = weighted_choice(choices)
				#Reduces number of nextMonomer by 1, since it is being used up in reaction, unless MAINTAIN is true
				if not MAINTAIN:
					monomerAmounts[nextMonomer - 1] -= 1
				#print("monomerAmounts: ", monomerAmounts)
				#Attaches next monomer to polymer chain
				polymer.append(nextMonomer)
				self.monomerUsageList.append(nextMonomer)
				index = 0
				for amount in monomerAmounts:
					self.monomerRemainingList[index].append(amount/ self.originalMonomerAmounts[index])
					index += 1
				self.monomerUsageSubList[startingMonomer - 1] += 1
			self.monomerUsageSubList = [float(i)/sum(self.monomerUsageSubList) for i in self.monomerUsageSubList]
				

		"-------------------------------------------------------------------------------------------------------------------"
		"***PROPAGATION STEP***"
		"-------------------------------------------------------------------------------------------------------------------"

		"Case for penultimate step (description will not be as in depth). For normal Mayo-Lewis case, scroll down."

		if PENULTIMATE:
			counter = 0
			while sum(monomerAmounts) > (1-self.conversion/100) * totalOriginalMonomerAmounts:	
				polymer = random.choice(self.polymerArray)
				choices = []
				#calculates weight chance for each monomer
				for monomerID in range(1, self.numMonomers + 1):
					#retrieving coefficient based on previous and currrent monomer, assumes hetergoenous penultimate monomer
					#assigning variables for clarity
					nextMonomer = monomerID - 1
					ultimateMonomer = polymer[-1] - 1
					penultimateMonomer = polymer[-2] - 1
					coeff = singleCoeffList[nextMonomer][ultimateMonomer][penultimateMonomer]
					# weight chance calulations for monomer attaching: coefficient*(amount of monomer remaining)
					chance = monomerAmounts[monomerID - 1] * coeff
					#adds a two element list to choices containing monomer and weight
					choices.append([monomerID, chance])
				#Using weighted_choice, selects next monomer
				try:
					nextMonomer = weighted_choice(choices)
				#If all weights are zero due to coefficients, then sort by relative amounts of monomer instead
				except AssertionError:
					monomerID = 1
					choices = []
					while monomerID <= self.numMonomers:
						choices.append([monomerID, monomerAmounts[monomerID - 1]])
						monomerID += 1
					nextMonomer = weighted_choice(choices)
				#Reduces number of nextMonomer by 1, since it is being used up in reaction, unless MAINTAIN is true
				if not MAINTAIN:
					monomerAmounts[nextMonomer - 1] -= 1
				#print("monomerAmounts: ", monomerAmounts)
				#Attaches next monomer to polymer chain
				polymer.append(nextMonomer)
				self.monomerUsageList.append(nextMonomer)
				index = 0
				for amount in monomerAmounts:
					self.monomerRemainingList[index].append(amount/ self.originalMonomerAmounts[index])
					index += 1
				self.monomerUsageSubList[nextMonomer - 1] += 1
				counter += 1
				if counter == self.numPolymers:
					counter = 0
					normalizedMonomerUsageSubList = [float(i)/sum(self.monomerUsageSubList) for i in self.monomerUsageSubList]
					for i in range(self.numMonomers):
						self.monomerUsageList[i].append(normalizedMonomerUsageSubList[i])
						self.monomerUsageSubList = [0]*self.numMonomers
			
				

		
		else:
			"***Propogation for standard Mayo-Lewis Case***"

			"The 'while' loop will add monomers to the chain until we reach we use up a percentage of the total starting monomers"
			"defined by user input into the 'Percent Conversion' box. Setting 'Percent Conversion' to 100% will continue to add "
			"monomers to growing chains until there are no remaining monomers, simulating a living polymerization."
			self.monomerUsageSubList = [0]*self.numMonomers
			counter = 0
			while sum(monomerAmounts) > (1-self.conversion/100) * totalOriginalMonomerAmounts:

				"Randomly choose a polymer chain to grow."
				"polymer, it appends one monomer to the growing chain, simulating an ideal homogenous polymer chain growth."
				polymer = random.choice(self.polymerArray)
				"For each polymer, iterate through all possible monomers which can be added. For each monomer, calculate"
				"the weight chance of the monomer to be added to the chain , defined as the product of the relevant rate "
				"constant 'k' times the number of monomers left unreacted 'f'"
				#A variable keeping track of current monomer
				choices = []
				for monomerID in range(1, self.numMonomers + 1):
					#Retrieveing coefficient based on previous and current monomer
					#get the last monomer on the growing chain
					terminatingMonomer = polymer[-1]
					#retrieve the relevant rate constant k
					k = singleCoeffList[terminatingMonomer - 1][monomerID - 1]
					# weight chance calulations for monomer attaching: coefficient*(amount of monomer remaining)
					chance = monomerAmounts[monomerID - 1] * k
					#print(monomerID, " amount of monomer remaining: ", monomerAmounts[monomerID - 1]);
					#Adds a two element list to choices containing monomer and weight
					choices.append([monomerID, chance])
					monomerID += 1
				#print ("currPolymerLength: ", currPolymerLength)
				#print("choices2: ", choices)
				#Using weighted_choice, selects next monomer
				try:
					nextMonomer = weighted_choice(choices)
				#If all weights are zero due to coefficients, then sort by relative amounts of monomer instead
				except AssertionError:
					monomerID = 1
					choices = []
					while monomerID <= self.numMonomers:
						choices.append([monomerID, monomerAmounts[monomerID - 1]])
						monomerID += 1
					nextMonomer = weighted_choice(choices)
				#Reduces number of nextMonomer by 1, since it is being used up in reaction, unless MAINTAIN is true
				if not MAINTAIN:
					monomerAmounts[nextMonomer - 1] -= 1
				#print("monomerAmounts: ", monomerAmounts)
				#Attaches next monomer to polymer chain
				polymer.append(nextMonomer)
				self.monomerUsageList.append(nextMonomer)
				index = 0
				for amount in monomerAmounts:
					self.monomerRemainingList[index].append(amount/ self.originalMonomerAmounts[index])
					index += 1
				self.monomerUsageSubList[nextMonomer - 1] += 1
				counter += 1
				if counter == self.numPolymers:
					counter = 0
					normalizedMonomerUsageSubList = [float(i)/sum(self.monomerUsageSubList) for i in self.monomerUsageSubList]
					for i in range(self.numMonomers):
						self.monomerUsageList[i].append(normalizedMonomerUsageSubList[i])
						self.monomerUsageSubList = [0]*self.numMonomers
			
				#print("currProgress: ", currProgress)

		"--------------------------------------------------------------------------------------------------------------------"
		"***OUTPUT RAW SIMULATION RESULTS INTO A TEXT FILE"
		"--------------------------------------------------------------------------------------------------------------------"

		text_file = open("polymerArray.txt", "w")
		json.dump(self.polymerArray, text_file)
		text_file.close()

		"--------------------------------------------------------------------------------------------------------------------"
		"***CALCULATE AND DISPLAY MONOMER COMPOSITION, WEIGHT AVG DP, NUMBER AVG DP, AND DISPERSITY INDEX***"
		"--------------------------------------------------------------------------------------------------------------------"

		self.compositionList = self.getComposition(self.polymerArray)
		numberAverageDP = round(self.numberAverageDP(self.polymerArray), 1)
		weightAverageDP = round(self.weightAverageDP(self.polymerArray, numberAverageDP), 1)
		dispersityIndex = round(self.disperityIndex(numberAverageDP, weightAverageDP), 3)
		if not self.compFrameExists:
			self.col3Sep = ttk.Separator(master = self.inputFrame, orient = Tk.VERTICAL)
			self.col3Sep.pack(side = Tk.LEFT, expand = True, fill = Tk.BOTH, pady = 1)
			self.compFrame = ttk.Frame(master = self.inputFrame)
			self.compFrame.pack(side = Tk.LEFT, padx = 0)
			self.compTkVarArray = []
			self.compLabelList = []
			for monomer in range(1, self.numMonomers + 1):
				dispCompFrame = ttk.Frame(master = self.compFrame)
				dispCompFrame.pack(side = Tk.TOP, pady = 1)
				if ALIAS:
					label = self.aliasList[monomer - 1]
				else:
					label = "Monomer " + str(monomer)
				fullLabel = label + " Composition: "
				compLabel = ttk.Label(master = dispCompFrame, text = fullLabel)
				self.compLabelList.append(compLabel)
				compLabel.pack(side = Tk.LEFT)
				compTkVar = Tk.DoubleVar()
				compTkVar.set(round(self.compositionList[monomer - 1], 2))
				self.compTkVarArray.append(compTkVar)
				compEntry = ttk.Entry(master = dispCompFrame, textvariable = compTkVar, width = 4)
				compEntry.pack(side = Tk.LEFT, padx = 5)
				self.compFrameExists = True
			numAvgDPFrame = ttk.Frame(master = self.compFrame)
			numAvgDPFrame.pack(side = Tk.TOP, pady = 1)
			label = "Number Average DP: "
			numAvgLabel = ttk.Label(master = numAvgDPFrame, text = label)
			numAvgLabel.pack(side = Tk.LEFT)
			self.numAvgTkVar = Tk.DoubleVar()
			self.numAvgTkVar.set(numberAverageDP)
			self.numAvgEntry = ttk.Entry(master = numAvgDPFrame, textvariable = self.numAvgTkVar, width = 5)
			self.numAvgEntry.pack(side = Tk.LEFT, padx = 5)

			weightAvgDPFrame = ttk.Frame(master = self.compFrame)
			weightAvgDPFrame.pack(side = Tk.TOP, pady = 1)
			label = "Weight Average DP: "
			weightAvgLabel = ttk.Label(master = weightAvgDPFrame, text = label)
			weightAvgLabel.pack(side = Tk.LEFT)
			self.weightAvgTkVar = Tk.DoubleVar()
			self.weightAvgTkVar.set(weightAverageDP)
			self.weightAvgEntry = ttk.Entry(master = weightAvgDPFrame, textvariable = self.weightAvgTkVar, width = 5)
			self.weightAvgEntry.pack(side = Tk.LEFT, padx = 5)

			dispersityFrame = ttk.Frame(master = self.compFrame)
			dispersityFrame.pack(side = Tk.TOP, pady = 1)
			label = "Dispersity Index: "
			dispersityLabel = ttk.Label(master = dispersityFrame, text = label)
			dispersityLabel.pack(side = Tk.LEFT)
			self.dispersityTkVar = Tk.DoubleVar()
			self.dispersityTkVar.set(dispersityIndex)
			self.dispersityEntry = ttk.Entry(master = dispersityFrame, textvariable = self.dispersityTkVar, width = 5)
			self.dispersityEntry.pack(side = Tk.LEFT, padx = 5)




		else:
			monomerID = 1
			for item in self.compTkVarArray:
				item.set(round(self.compositionList[monomerID - 1], 2))
				monomerID += 1
			self.numAvgTkVar.set(numberAverageDP)
			self.weightAvgTkVar.set(weightAverageDP)
			self.dispersityTkVar.set(dispersityIndex)


		"***Visualize Polymers***"
		self.visualizePolymers(self.polymerArray)

		"***Plot Polymer Data***"
		self.plotCompositions(False)
		center(root)

		"***Unlock***"
		self.simulateLocked = False
	
	"----------------------------------------------------------------------------------------------------------------------------"
	"***PLOT THE USER-SELECTED GRAPHS***"
	"----------------------------------------------------------------------------------------------------------------------------"

	def plotCompositions(self, update):
		try:
			polymerArray = self.polymerArray
		except:
			errorMessage("Please simulate first!", 300)
			return
		try:
			self.graph1Type = self.graphType1TkVar.get()
			self.graph2Type = self.graphType2TkVar.get()
			if self.graph1Type == "Hydrophobic Blocks" or self.graph1Type == "Hydrophilic Blocks" or self.graph2Type == "Hydrophobic Blocks" or self.graph2Type == "Hydrophilic Blocks":
					for phobicity in self.hphobList:
						if phobicity == "None":
							raise phobicityNotSpecified("hi")
		except phobicityNotSpecified:
			errorMessage("Please specficy hydrophobicities in the Options Tab.", 330)
			return
		#destroys canvas if necessary
		#style to use
		style.use('classic')
		style.use(STYLE)
		#retrieving graphType variables
		self.graph1Type = self.graphType1TkVar.get()
		self.graph2Type = self.graphType2TkVar.get()
		print("graph1type: ", self.graph1Type)
		#Plot and Figure formatting
		if not self.canvasExists:
			self.plotFigure = plt.figure(figsize=(5.5, 3.3),dpi =100)
		if self.canvasExists:
			if self.subplot1Exists:
				self.subplot1.clear()
			if self.subplot2Exists:
				self.subplot2.clear()
		if self.graph1Type == "None" and self.graph2Type == "None":
			return
		if self.graph1Type == "None":
			if not self.canvasExists:
				self.subplot1 = self.plotFigure.add_subplot(111)
			else:
				if self.subplot2Exists:
					self.plotFigure.delaxes(self.subplot1)
					self.plotFigure.delaxes(self.subplot2)
					self.subplot1 = self.plotFigure.add_subplot(111)
			self.subplot1.set_color_cycle(COLORARRAY)
			self.subplot1.tick_params(labelsize = 7)
			self.subplot1Exists = True
			self.subplot2Exists = False
			self.graphSubPlot(polymerArray, self.graph2Type, self.subplot1, 1)
		elif self.graph2Type == "None":
			if not self.canvasExists:
				self.subplot1 = self.plotFigure.add_subplot(111)
			else:
				if self.subplot2Exists:
					self.plotFigure.delaxes(self.subplot1)
					self.plotFigure.delaxes(self.subplot2)
					self.subplot1 = self.plotFigure.add_subplot(111)
			self.subplot1.set_color_cycle(COLORARRAY)
			self.subplot1.tick_params(labelsize = 7)
			self.subplot1Exists = True
			self.subplot2Exists = False
			self.graphSubPlot(polymerArray, self.graph1Type, self.subplot1, 1)
		else:
			if not self.canvasExists:
				self.subplot1 = self.plotFigure.add_subplot(121)
				self.subplot2 = self.plotFigure.add_subplot(122)
				self.subplot1.tick_params(labelsize = 7)
				self.subplot2.tick_params(labelsize = 7)
			else:
				if not self.subplot1Exists:
					self.subplot1 = self.plotFigure.add_subplot(121)
					self.subplot1.tick_params(labelsize = 7)
				if not self.subplot2Exists:
					self.subplot2 = self.plotFigure.add_subplot(121)
					self.subplot2.tick_params(labelsize = 7)
			self.subplot1.set_color_cycle(COLORARRAY)
			self.subplot2.set_color_cycle(COLORARRAY)
			self.subplot1Exists = True
			self.subplot2Exists = True
			self.graphSubPlot(polymerArray, self.graph1Type, self.subplot1, 1)
			self.graphSubPlot(polymerArray, self.graph2Type, self.subplot2, 2)	
			#leg = self.subplot1.legend()
			#leg.draggable()
		# A tk.DrawingArea
		#imbedding matplotlib graph onto canvas
		if not self.canvasExists:
			self.canvas = FigureCanvasTkAgg(self.plotFigure, master = root)
		else:
			self.canvas.draw()
		self.canvas.show()
		#Imbedding matplotlib toolbar onto canvas
		if not self.canvasExists:
			self.toolbar = NavigationToolbar2TkAgg(self.canvas, root)
			self.toolbar.update()
			self.canvas._tkcanvas.pack(side = Tk.BOTTOM, fill = Tk.BOTH, expand = 1)
			self.canvas.get_tk_widget().pack(side = Tk.BOTTOM, fill = Tk.BOTH, expand = 1)
		self.canvasExists = True
		self.plotted = True
		#w = root.winfo_screenwidth()
		#h = root.winfo_screenheight()
		#root.geometry("%dx%d+0+0" % (w, h))
		return
		#very bad fix for draggable, should correct
		#self.plotCompositions(True)

	"-------------------------------------------------------------------------------------------------------------------------"
	"***VISUALIZE THE POLYMERS WITH COLORED SQUARES***"
	"-------------------------------------------------------------------------------------------------------------------------"

	def visualizePolymers(self, polymerArray):
		if DYAD:
			polymerArrayToUse = self.getDyad(polymerArray)
			monomerRange = self.numMonomers * 2 + 1
		else:
			polymerArrayToUse = polymerArray
			monomerRange = self.numMonomers + 1
		if not self.visualizationFrameExists:
			#LabelFrame for visualizeCanvas
			self.visualizationFrame = ttk.LabelFrame(master = root, text = "Polymer Visualization")
			self.visualizationFrame.pack(side = Tk.BOTTOM, fill = Tk.BOTH, expand = 0, padx = 7, pady = 0)
		if self.visualizationFrameExists:
			self.visualizeCanvas.destroy()
		#update visuals to get correct sizing
		self.visualizationFrame.update()
		#variable to keep track of frame width
		self.visualFrameWidth = self.visualizationFrame.winfo_width()
		#print("Width", self.visualFrameWidth)
		numRows = self.numPolyToShow.get()
		maxPolymerLength = len(max(self.polymerArray[:numRows], key=lambda x: len(x)))
		if numRows > self.numPolymers * self.numSimulations: 
			numRows = self.numPolymers * self.numSimulations
		#parameters for canvas height and width
		canvasHeight = 130
		canvasWidth = self.visualFrameWidth

		#Maximizes size of squares
		if (canvasHeight - 25) / numRows <= (canvasWidth - 40) / maxPolymerLength:
			size = (canvasHeight - 25) / numRows
			canvasWidth = maxPolymerLength * size + 20
		else:
			size = (canvasWidth - 40) / maxPolymerLength
			canvasHeight = numRows * size + 10
		self.visualFrameWidth = self.visualizationFrame.winfo_height()
		#Canvas for visualization
		self.visualizeCanvas = Tk.Canvas(master = self.visualizationFrame, width = canvasWidth, height = canvasHeight)
		self.visualizeCanvas.pack()
		self.polymerImage = Image.new("RGB", (int(size*self.polymerLength) + 1, int(size*numRows) + 1), "white")
		draw = ImageDraw.Draw(self.polymerImage)
		#colors
		#line colors to use
		#COLORARRAY = ['#4D4D4D','#5DA5DA', '#F15854', '#DECF3F', '#60BD68', '#F17CB0', '#B276B2', '#FAA43A']
		#Pad Parameters

		ulx = 20
		uly = 10
		ulx2 = 0
		uly2 = 0
		#Visualizes polymers, number of polymers visualized based on numRows
		for row in range(0, numRows):
			#iterates through an array representation of monomer and adds a square with corresponding color
			for monomer in polymerArrayToUse[row]:
				color = COLORARRAY[monomer - 1]
				dcolor = DCOLORARRAY[monomer - 1]
				self.visualizeCanvas.create_rectangle(ulx, uly + size * row, ulx + size, uly + size * (row + 1), fill = color, activefill = dcolor)
				draw.rectangle(((ulx2, uly2 + size * row), (ulx2 + size, uly2 + size * (row + 1))), fill = color, outline = "black")
				ulx += size
				ulx2 += size
			ulx = 20
			ulx2 = 0
		self.visualizationFrameExists = True
		self.polymerImage.save("polymerImage.jpg")

	"--------------------------------------------------------------------------------------------------------------------------------"
	"***FUNCTIONS FOR PARSING USER INPUTTED DATA***"
	"--------------------------------------------------------------------------------------------------------------------------------"


	"***Retrieves monomer ratio data***"
	#Converts an array of ttk.Entrys of ratios into an int array of starting monomer amounts
	#Note: might have result in total amount fo momoners being slightly more tcanvashan inital total monomers due to ceiling divide
	def getMonomerAmounts(self):
		#A list of starting monomer amounts
		monomerAmounts = []
		#A list of starting ratios (not tkvars!)
		startingWeightList = []
		#Converts a list of Tkvars to a list of floats
		for entry in self.startingRatiosTkList:
			assert(float(entry.get()) > 0)
			startingWeightList.append(float(entry.get()))
		totalWeight = sum(startingWeightList)
		for weight in startingWeightList:
			#calculates number of monomers from monomer ratio and total monomers, ceiling dividing
			numMonomers = math.ceil(self.totalMonomers * weight / totalWeight)
			monomerAmounts.append(numMonomers)
			assert(weight > 0)
		#print("monomerAmounts: ", monomerAmounts)
		return monomerAmounts

	"***Retrieves standard reactivity ratio data***"
	#Converts a 2D array of ttk.Entrys for coefficients into a 2D double array
	def getCoefficients(self):
		coeffList = []
		for entry in self.coefficientList:
			if entry == 0:
				coeffList.append(0)
			else:
				singleCoeffList = []
				coeffList.append(singleCoeffList)
				for innerEntry in entry: 
					if innerEntry == 0:
						singleCoeffList.append(0)
					elif ',' in list(innerEntry.get()):
						raise notInEuropeError("You're in America!")
					else:
						assert(float(innerEntry.get()) >= 0)
						singleCoeffList.append(float(innerEntry.get()))
		print("coeffList: ", coeffList)
		return coeffList

	def getReactivity(self, coeffList, numMonomers):
		if numMonomers == 2:
			return coeffList
		else:
			#rrSubList: a list of reactivity ratios for a single monomer
			#>>>print(coeffList)
			#[[1.0, 1.0], [1.0, 1.0], [1.0, 1.0]]
			reactivityList = []
			for currMonomer in range(1, numMonomers + 1):
				rrSubList = coeffList[currMonomer - 1]
				reactivitySubList = []
				monomerAccess = 0
				for secondMonomer in range(1, numMonomers + 1):
					if secondMonomer == currMonomer:
						reactivitySubList.append(1)
					else:
						reactivity = 1/rrSubList[monomerAccess]
						reactivitySubList.append(reactivity)
						monomerAccess += 1
				reactivityList.append(reactivitySubList)
			print("reactivityList: ", reactivityList)
			return reactivityList




	"***Retrieves penultimate reactivity ratio data"
	def getPenultimateCoeff(self):
		coeffList = []
		for nextMonomerList in self.coefficientList:
			ultimateList = []
			coeffList.append(ultimateList)
			for ultimateMonomerList in nextMonomerList:
				penultimateCoeffList = []
				ultimateList.append(penultimateCoeffList)
				for penultimateCoeffEntry in ultimateMonomerList:
					if ',' in list(penultimateCoeffEntry.get()):
						raise notInEuropeError("You're in America!")
					else:
						assert(float(penultimateCoeffEntry.get()) >= 0)
						penultimateCoeffList.append(float(penultimateCoeffEntry.get()))
		print("coeffList: ", coeffList)
		return coeffList

	"----------------------------------------------------------------------------------------------------------------------------"
	"***FUNCTIONS FOR ANALYZING SIMULATED POLYMER ARRAY DATA"
	"----------------------------------------------------------------------------------------------------------------------------"
	"***Analysis for Total Mass***"
	def totalMass(self, polymerArray):
		totalMass = 0
		for polymer in polymerArray:
			totalMass += len(polymer)
		return totalMass

	"***Analysis for Number Average***"
	#Given a list of polymers, returns the number average DP
	def numberAverageDP(self, polymerArray):
		totalMass = self.totalMass(polymerArray)
		numberAverageDP = totalMass / self.numPolymers
		return numberAverageDP

	#Given a list of polymers, return the weight average DP
	"***Analysis for Weight Average***"
	def weightAverageDP(self, polymerArray, numberAverageDP):
		weightAverageDP = 0
		for polymer in polymerArray:
			M = len(polymer)
			weightAverageDP += M*M
		weightAverageDP = weightAverageDP / numberAverageDP / self.numPolymers
		return weightAverageDP

	#Given number and weight average DP, return the dispersity index
	"***Analysis for Dispersity Index***"
	def disperityIndex(self, numberAverageDP, weightAverageDP):
		return weightAverageDP / numberAverageDP 

	"***Analysis for Run Length***"
	#returns an array of numbers for each consecutive monomer; to be used in histogram plotting
	def getHistogramData(self, polymerArray, monomerID, indexLimit):
		#counting number of monomers in polymerArray
		#flat_list = [item for sublist in polymerArray for item in sublist]
		#print('number of monomers used: ' , len(flat_list))
		#print('indexLimit: ', indexLimit)
		#initializing array to be returned
		histogramData = []
		#count through all polymers
		for polymer in polymerArray: 
			#constant to keep track of polymer index, to know when to stop counting
			polymerIndex = 1
			numConsecutive = 0
			#iterate through each polymer until monomerLimit is hit and count consecutive polymers
			for monomer in polymer:
				#if index limit is reached, add any consecutives to histogram, and break from for loop
				if polymerIndex >= indexLimit:
					if  monomer == monomerID:
						numConsecutive += 1
					#print(numConsecutive)
					if numConsecutive > 0:
						count = 0
						while count < numConsecutive:
							histogramData.append(numConsecutive)
							count += 1
					break
				#if monomer is not consecutive and monomer before was, add consecutive number to data and reset values
				if monomer != monomerID and numConsecutive > 0:
					count = 0
					while count < numConsecutive:
						histogramData.append(numConsecutive)
						count += 1
					numConsecutive = 0
					#polymerIndex += 1
					#continue
				#increment consecutive counter by 1 if monomer is consecutive
				if  monomer == monomerID:
					numConsecutive += 1
					#polymerIndex += 1
					#continue
				polymerIndex += 1
				#print(polymerIndex)
		#print("length histogramData: ", len(histogramData))
		if not histogramData:
			histogramData = [0]
		return histogramData

	"***Analysis for Dyads***"
	#returns an array, same size as polymerArray, with dyad monomer being assign different numbers
	def getDyad(self, polymerArray):
		dyadArray = []
		for polymer in polymerArray:
			polymerDyad = []
			for monomer in polymer:
				if polymerDyad == []:
					polymerDyad.append(monomer)
					prevMonomer = monomer
					continue
				if monomer == prevMonomer:
					polymerDyad[-1] = prevMonomer + self.numMonomers
					polymerDyad.append(prevMonomer + self.numMonomers)
				else:
					polymerDyad.append(monomer)
					prevMonomer = monomer
			dyadArray.append(polymerDyad)
		return dyadArray

	"***Monomer composition analysis***"
	#Returns a list of the percent compostion of each monomer, in monomerID order
	def getComposition(self, polymerArray):
		compositionList = [0] * self.numMonomers
		for monomerID in range(1, self.numMonomers + 1):
			for polymer in polymerArray:
				for monomer in polymer:
					if monomer == monomerID:
						compositionList[monomerID - 1] += 1
		totalMonomers = sum(compositionList)
		#print("precomp: ", compositionList)
		#print("totalMonomers: ", totalMonomers)
		for monomerID in range(1, self.numMonomers + 1):
			compositionList[monomerID - 1] = compositionList[monomerID - 1] / totalMonomers
		#print("compositionList: ", compositionList)
		return compositionList

	"***Polymer composition analysis, DEPRECATED***"	
	#Returns a list of the percent composition of each monomer at each index, in monomerID order
	def getFullCompositionAtIndex(self, polymerArray):
		fullCompList = []
		for monomerID in range(1, self.numMonomers + 1):
			compositionList = [0] * self.numMonomers
			monomerCompList = []
			for monomerIndex in range(0, self.polymerLength):
				for monomerID2 in range(1, self.numMonomers + 1):
					for polymer in polymerArray:
						if polymer[monomerIndex] == monomerID2:
							compositionList[monomerID2 - 1] += 1
				totalMonomers = sum(compositionList)
				monomerCompList.append(compositionList[monomerID - 1] / totalMonomers)
			fullCompList.append(monomerCompList)
			#totalMonomers = sum(compositionList)
		#print("precomp: ", compositionList)
		#print("totalMonomers: ", totalMonomers)
		#print("index: ", index)
		return fullCompList
	"***DP Distribution Analysis***"
	def DP_Distribution(self, polymerArray):
		DP_Distribution = []
		for polymer in polymerArray:
			DP_Distribution.append(len(polymer))
		return DP_Distribution

	"***Hydrophobic/Hydrophillic analysis***"
	#takes in a polymerArray, an converts monomer number to 0 for hydrophobic, 1 for hydrophilic
	def convertHydro(self, polymerArray):
		print("hphobList: ", self.hphobList)
		newPolymerArray = []
		for polymer in polymerArray:
			newPolymer = []
			for monomer in polymer:
				#print(monomer)
				if self.hphobList[monomer - 1] == "Hydrophobic":
					newPolymer.append(0)
				elif self.hphobList[monomer - 1] == "Hydrophilic":
					newPolymer.append(1)
			newPolymerArray.append(newPolymer)
		return newPolymerArray

	"***Alter histogram data for histogram plotting***"
	def getHistPlotvals(self, histogramMonomer):
			histogramNumberLimit = int(self.polymerLength)
			print("histogramLimit: ", histogramNumberLimit)
			histDataList = []
			for monomerID in range(1, self.numMonomers + 1):
				histDataList.append(self.getHistogramData(self.polymerArray, monomerID, histogramNumberLimit))
			histogramData = self.getHistogramData(self.polymerArray, histogramMonomer, histogramNumberLimit)
			maxList = []
			#finding maximum 
			for data in histDataList:
				maxList.append(max(data))
			maximum = max(maxList)
			print('maximum: ', maximum)
			#print(histogramData)

			binwidth = 1
			hist, bins = np.histogram(histogramData, bins=range(1, maximum + binwidth + 1, binwidth))
			# print("hist: ", hist)
			# print("bins: ", bins)
			# print("sum hist: ", sum(hist))
			# print("normFactor: ", self.polymerLength * self.numPolymers)
			#first normalization step
			hist = hist / (self.polymerLength * self.numPolymers) 
			print("newHist: ", hist)
			return [bins, hist, histogramData]

	"----------------------------------------------------------------------------------------------------------------------------"
	"***GRAPHING FUNCTIONS"
	"----------------------------------------------------------------------------------------------------------------------------"
	
	def graphSubPlot(self, polymerArray, graphType, subplot, number):
		if graphType == "Monomer Usage" or graphType == "Monomer Occurrences" or graphType == "Polymer Compositions":
			if graphType == "Monomer Occurrences":
				if DYAD:
					polymerArrayToUse = self.getDyad(polymerArray)
					monomerRange = self.numMonomers * 2 + 1
				else:
					polymerArrayToUse = polymerArray
					monomerRange = self.numMonomers + 1
				#Iterates through each unique monomer and plots composition
				for monomer in range(1, monomerRange):
					#x-axis array
					polymerIndex = list(range(1, self.polymerLength + 1))
					#y-axis array initation
					monomercounts = [0] * self.polymerLength
					#inputs counts into y-axis array
					#adjust axis title
					subplot.set_ylabel("Normalized Monomer Occurrences", labelpad=5, fontsize = 9)
					varLengths = True
					if not varLengths:
						for index in polymerIndex:
							count = 0
							for polymer in polymerArrayToUse:
								if polymer[index - 1] == monomer:
									count += 1
							monomercounts[index - 1] = float(float(count) / float(self.numSimulations) / float(self.numPolymers))
					if monomer > self.numMonomers:
						if ALIAS:
							label = self.aliasList[monomer - self.numMonomers - 1] + " Homodyad"
						else:
							label = "Homodyad " + str(monomer - self.numMonomers)
					else:
						if ALIAS:
							label = self.aliasList[monomer - 1]
						else:
							label = "Monomer " + str(monomer)
					totalOriginalMonomerAmounts = sum(list(self.originalMonomerAmounts))
					monomerUsage = self.monomerUsageList[monomer - 1]
					#print("monomerUsage: ", monomerUsage)
					conversionIndex = range(len(monomerUsage))
					conversionIndex = [i*self.numPolymers/totalOriginalMonomerAmounts*100 for i in conversionIndex]
					#polymerIndex = [num * 100 / self.raftRatio for num in polymerIndex]
					#polymerIndex = range(len(self.monomerUsageData))
					curve = subplot.plot(conversionIndex, monomerUsage, label = label)
					#setting axis limits
					subplot.set_ylim([0,1])
			#graphs Percentage of Monomer Remaining
			if graphType == "Monomer Usage":
				for monomer in range(1, self.numMonomers + 1):
					#x-axis array
					polymerIndex = list(range(1, self.polymerLength + 1))

					#y-axis array initation
					monomercounts = [0] * self.polymerLength
					#adjust axis title
					subplot.set_ylabel("Monomer Left (M/Mo)", labelpad=5, fontsize = 9)
					#adjust y axis limiys
					subplot.set_ylim([0,1])
					#variable to keep track of average number of monomers consumed
					# monomersConsumed = 0
					# for index in polymerIndex:
					# 	count = 0
					# 	for polymer in polymerArray:
					# 		if polymer[index - 1] == monomer:
					# 			count += 1
					# 	startingMonomerAmount = self.originalMonomerAmounts[monomer - 1]
					# 	#calculates monomer consumed
					# 	monomersConsumed += count / self.numSimulations
					# 	#calculated percentage of monomer remaining
					# 	percentageRemaining = (startingMonomerAmount - monomersConsumed) / startingMonomerAmount
					# 	monomercounts[index - 1] = percentageRemaining
					# polymerIndex = [num * 100 / self.raftRatio for num in polymerIndex]
				#debugging purposes
				#print(polymerIndex)
				#print(monomercounts)
				#plots x and y arrays
					monomerRemaining = self.monomerRemainingList[monomer -1]
					conversionIndex = range(len(monomerRemaining))
					totalOriginalMonomerAmounts = sum(list(self.originalMonomerAmounts))
					conversion = [i/totalOriginalMonomerAmounts*100 for i in conversionIndex]

					if ALIAS:
						labelToUse = self.aliasList[monomer - 1]
						curve = subplot.plot(conversion, monomerRemaining, label = labelToUse)
					else:
						curve = subplot.plot(conversion, monomerRemaining, label = "Monomer " + str(monomer))
			#legend-screw matplotlib; so fucking hard to format
			handles, labels = subplot.get_legend_handles_labels()
			if LEGEND:
				lgd = subplot.legend(handles, labels, prop = {'size':7}, loc = "best")
				lgd.draggable(state = True)
			subplot.set_xlabel("Conversion", labelpad = 0, fontsize = 9)
		elif graphType == "DP Distribution":
			#adjust axis title
			DP_Distribution = self.DP_Distribution(self.polymerArray)
			#list of data for all monomers
			maximum = max(DP_Distribution)
			minimum = min(DP_Distribution)
			binwidth = 1
			hist, bins = np.histogram(DP_Distribution, bins=range(minimum, maximum + binwidth + 1, binwidth))
			widths = np.diff(bins)
			subplot.bar(bins[:-1], hist, widths)
			#subplot.set_xticks(arange(min(DP_Distribution), max(DP_Distribution) + binwidth, binwidth))
			subplot.hist(DP_Distribution, bins)
			#plt.xticks(rotation = 45)
			subplot.set_xlabel("DP Distribution", labelpad=0, fontsize = 9)
		elif graphType == "Run Length":
			#obtain histogram limit
			try:
				self.histogramLimit = float(self.histogramLimitTkVar.get())
				assert(self.histogramLimit <= 1)
				assert(self.histogramLimit > 0)
			except AssertionError:
				errorMessage("Percent to Analyze Value Invalid!", 220)
				return
			#retrieving all needed variables from inputs
			if number == 1:
				histogramMonomer = int(self.histogramMonomer1TkVar.get())
			if number == 2:
				histogramMonomer = int(self.histogramMonomer2TkVar.get())
			#get all the histogram data so can set max for histogram
				#print("histafter: ", weightedHist)
			histPlotVals = self.getHistPlotvals(histogramMonomer)
			bins = histPlotVals[0]
			hist = histPlotVals[1]
			histogramData = histPlotVals[2]
			widths = np.diff(bins)
			subplot.bar(bins[:-1], hist, widths, color = COLORARRAY[histogramMonomer - 1])
			#subplot.hist(histogramData, bins=range(min(histogramData), max(histogramData) + binwidth, binwidth),
			 #color = COLORARRAY[histogramMonomer - 1], normed = True)
			if graphType == "Run Length":
				subplot.set_xlabel("Monomer %i Run Length" %histogramMonomer, labelpad = 0, fontsize = 9)
				subplot.set_ylabel("Normalized Counts", labelpad=5, fontsize = 9)
				#subplot.set_ylim([0,yUpperLim])
			elif graphType == "Block Size":
				subplot.set_xlabel("Monomer %i Block Size" %histogramMonomer, labelpad = 0, fontsize = 9)
				subplot.set_ylabel("Normalized Count", labelpad=5, fontsize = 9)
				#subplot.set_ylim([0,yUpperLim])
			subplot.set_xticks(arange(min(histogramData), max(histogramData) + 1, 1))
			#print(min(histogramData))
			#print(max(histogramData))
		elif graphType == "Hydrophobic Blocks" or graphType == "Hydrophilic Blocks":
			if graphType == "Hydrophobic Blocks":
				phobToShow = 0
			else:
				phobToShow = 1
			polymerArray = self.convertHydro(polymerArray)
			#print(polymerArray)
			#obtain histogram limit
			histogramNumberLimit = int(self.polymerLength)
			#get all the histogram data so can set max for histogram
			histDataList = []
			for hphobicity in range(0,2):
				histDataList.append(self.getHistogramData(polymerArray, hphobicity, histogramNumberLimit))
			histogramData = self.getHistogramData(polymerArray, phobToShow, histogramNumberLimit)
			maxList = []
			#print("histDataList: ", histDataList)
			#finding maximum 
			for data in histDataList:
				if data:
					maxList.append(max(data))
			maximum = max(maxList)
			#print('maximum: ', maximum)
			#print(histogramData)
			binwidth = 1
			hist, bins = np.histogram(histogramData, bins=range(1, maximum + binwidth + 1, binwidth))
			#first normalization step
			hist = hist / (self.polymerLength * self.numPolymers) 
				#print("histafter: ", weightedHist)
			widths = np.diff(bins)
			subplot.bar(bins[:-1], hist, widths, color = COLORARRAY[phobToShow])
			if phobToShow == 0:
				subplot.set_xlabel("Hydrophobic Block Size", labelpad = 0, fontsize = 9)
			if phobToShow == 1:
				subplot.set_xlabel("Hydrophilic Block Size", labelpad = 0, fontsize = 9)
			subplot.set_ylabel("Normalized Separation", labelpad=5, fontsize = 9)
			if histogramData:
				minimum = min(histogramData)
				maximum = max(histogramData)
			else:
				minimum = 0
				maximum = 0
			subplot.set_xticks(arange(minimum, maximum + 1, 1))

	"------------------------------------------------------------------------------------------------------------------------------"
	"***SAVING STATE***"
	"------------------------------------------------------------------------------------------------------------------------------"
	
	def saveState(self):
		ratiosList = []
		#Asserts that there are no input errors. Shows errorMessage if error caught.test
		try:
			self.totalMonomers = int(self.totalMonomersTkVar.get())
			monomerAmounts = self.getMonomerAmounts()
			if PENULTIMATE:
				singleCoeffList = self.getPenultimateCoeff()
			else:
				singleCoeffList = self.getCoefficients()
			self.numSimulations = NUM_SIMULATIONS
			self.raftRatio = float(self.raftRatioTkVar.get())
			self.histogramLimit = float(self.histogramLimitTkVar.get())
			#Converts a list of Tkvars to a list of floats
			for entry in self.startingRatiosTkList:
				assert(float(entry.get()) > 0)
				ratiosList.append(float(entry.get()))
			assert(self.histogramLimit <= 1)
			assert(self.histogramLimit > 0)
			assert(self.totalMonomers > 0)
			assert(self.numSimulations > 0)
			assert(self.raftRatio > 0)
			assert(self.totalMonomers * self.numSimulations <= MONOMER_CAP)
		except ValueError:
			errorMessage("Please input valid parameters!", 220)
			return
		except notInEuropeError:
			errorMessage("You are not in Europe!", 220)
			return
		except AssertionError:
			errorMessage("Please input valid parameters!", 220)
			return
		startingWeightList = []
		#error checking to see if config file exists
		# checking to see what state to save into
		stateNumber = 1
		while True:
			if os.path.exists("state%i.txt" %(stateNumber)):
				#errorMessage("config.txt does not exist!", 220)
				stateNumber += 1
				continue
			else:
				break
		print("stateNumber: ", stateNumber)
		print("singleCoeffList: ", singleCoeffList)
		nextSetting = LAST_SETTING + 1
		file = open("state%i.txt" %(stateNumber), "w")
		file.write("\n# Setting %i \nNumber of Unique Monomers = %i " %(nextSetting, self.numMonomers))
		monomerIndex = 1
		#Writing starting ratios
		while monomerIndex <= self.numMonomers:
			file.write("\nMonomer %i Ratio = %i" %(monomerIndex, ratiosList[monomerIndex - 1]))
			monomerIndex += 1
		monomerIndex = 1
		print("single:", singleCoeffList)
		#writing coefficients for penultimate model
		if PENULTIMATE:
			file.write("\nPenultimate")
			for nextMonomer in range(1, self.numMonomers + 1):
				for ultimate in range(1, self.numMonomers + 1):
					for penultimate in range(1, self.numMonomers + 1):
						coeff = singleCoeffList[nextMonomer - 1][ultimate - 1][penultimate - 1]
						file.write("\n%i-%i-%i = %f" %(penultimate, ultimate, nextMonomer, coeff))
			file.write("\nend")
			file.close()
		#writing coefficients for normal model
		else:
			file.write("\nStandard")
			while monomerIndex <= self.numMonomers:
				rrIndex = 0
				for secondMonomer in range(1, self.numMonomers + 1):
					if self.numMonomers == 2:
						coefflist = singleCoeffList[monomerIndex - 1]
						file.write("\n%i-%i = %f" %(monomerIndex, secondMonomer, (coefflist[secondMonomer - 1])))
						rrIndex += 1	
					else:
						if secondMonomer != monomerIndex:
							coefflist = singleCoeffList[monomerIndex - 1]
							print("coefflist: ", coefflist)
							file.write("\n%i-%i = %f" %(monomerIndex, secondMonomer, (coefflist[rrIndex])))
							rrIndex += 1
				monomerIndex += 1
			file.write("\nend")
			file.close()
		global LAST_SETTING
		LAST_SETTING += 1
		infoMessage("Save Successful", "State successfully saved into state %i!" %stateNumber, 300)
		#self.saveButton.update()
		return

	"---------------------------------------------------------------------------------------------------------------------------"
	"***'OPTIONS' DISPLAY****"
	"---------------------------------------------------------------------------------------------------------------------------"

	#displays toplevel options window
	def displayOptions(self):
		self.optionsWindow = Tk.Toplevel()
		self.optionsWindow.grab_set()
		self.optionsWindow.focus_force()
		self.optionsWindow.title("Options")
		self.allOptionsFrame = ttk.Frame(master = self.optionsWindow)
		self.allOptionsFrame.pack(side = Tk.TOP, pady = 5, padx = 5)
		self.options1Frame = ttk.Frame(master = self.allOptionsFrame)
		self.options1Frame.pack(side = Tk.LEFT, padx = 5, pady = 5)
		self.styleFrame = ttk.Frame(master = self.options1Frame)
		self.styleFrame.pack(side = Tk.TOP)
		self.styleLabel = ttk.Label(master = self.styleFrame, text = "Style: ")
		self.styleLabel.pack(side = Tk.LEFT)
		self.styleTkVar = Tk.StringVar()
		self.styleComboBox = ttk.Combobox(master = self.styleFrame, values = ["bmh", "classic", "dark_background", "fivethirtyeight", "ggplot", 
			"grayscale", "seaborn-colorblind", "seaborn-dark", "seaborn-dark-pallete", "seaborn-darkgrid", "seaborn-deep", "seaborn-muted", "seaborn-notebook",
			"seaborn-paper", "seaborn-pastel", "seaborn-poster", "seaborn-talk", "seaborn-ticks", "seaborn-white", "seaborn-whitegrid"], 
		state = "readonly", textvariable = self.styleTkVar)
		self.styleTkVar.set(STYLE)
		self.styleComboBox.pack(side = Tk.LEFT)
		#Frame for dyad
		self.dyadFrame = ttk.Frame(master = self.options1Frame)
		#DEPRACTES DYAD
		#self.dyadFrame.pack(side = Tk.TOP, pady = 3)
		#label for dyad
		self.dyadLabel = ttk.Label(master = self.dyadFrame, text = "Enable Homodyad Detection:")
		self.dyadLabel.pack(side = Tk.LEFT)
		#checkbox for dyad
		self.dyadCheckButton = ttk.Checkbutton(master = self.dyadFrame, text = None)
		self.dyadCheckButton.pack(side = Tk.LEFT, padx = 5)
		#Setting penultimateTkVar to PENULTIMATE global variable
		self.dyadTkVar = Tk.IntVar()
		self.dyadCheckButton["variable"] = self.dyadTkVar
		self.dyadTkVar.set(DYAD)
		self.initialSetFrame = ttk.Frame(master = self.options1Frame)
		self.initialSetFrame.pack(side = Tk.TOP)
		self.initialSetLabel = ttk.Label(master = self.initialSetFrame, text = "Initial Monomer: ")
		self.initialSetLabel.pack(side = Tk.LEFT)
		self.initialSetTkVar = Tk.StringVar()
		self.initialSetOptions = ["Weighted"]
		for monomerID in range(1, self.numMonomers + 1):
			self.initialSetOptions.append("Monomer %i" %monomerID)
		self.initialSetComboBox = ttk.Combobox(master = self.initialSetFrame, values = self.initialSetOptions, state = "readonly", 
			textvariable = self.initialSetTkVar, width = 11)
		self.initialSetTkVar.set(self.initialSetOptions[SETINITIAL])
		self.initialSetComboBox.pack(side = Tk.LEFT)
		#Frame for numSimulations label and Entry
		self.numSimsFrame = ttk.Frame(master = self.options1Frame)
		self.numSimsFrame.pack(side = Tk.TOP)
		#Label for numSims Entry
		self.numSimsLabel = ttk.Label(master = self.numSimsFrame, text = "Number of Simulations:   ")
		self.numSimsLabel.pack(side = Tk.LEFT, pady = 3)
		#Entry for numSimulations
		self.numSimsEntry = ttk.Entry(master = self.numSimsFrame, width = 3)
		self.numSimsEntry.pack(side = Tk.LEFT, padx = 3, pady = 3)
		#Setting number of simulations to 1000
		self.numSimsTkVar = Tk.IntVar()
		self.numSimsEntry["textvariable"] = self.numSimsTkVar
		self.numSimsTkVar.set(NUM_SIMULATIONS)
		self.maintainFrame = ttk.Frame(master = self.options1Frame)
		self.maintainFrame.pack(side = Tk.TOP)
		self.maintainLabel = ttk.Label(master = self.maintainFrame, text = "Maintain Composition")
		self.maintainLabel.pack(side = Tk.LEFT)
		self.maintainCheckButton = ttk.Checkbutton(master = self.maintainFrame, text = None)
		self.maintainCheckButton.pack(side = Tk.LEFT, padx = 5)
		self.maintainTkVar = Tk.IntVar()
		self.maintainCheckButton["variable"] = self.maintainTkVar
		self.maintainTkVar.set(MAINTAIN)
		self.legendFrame = ttk.Frame(master = self.options1Frame)
		self.legendFrame.pack(side = Tk.TOP)
		self.legendLabel = ttk.Label(master = self.legendFrame, text = "Enable Legend")
		self.legendLabel.pack(side = Tk.LEFT)
		self.legendCheckButton = ttk.Checkbutton(master = self.legendFrame, text = None)
		self.legendCheckButton.pack(side = Tk.LEFT, padx = 5)
		self.legendTkVar = Tk.IntVar()
		self.legendCheckButton["variable"] = self.legendTkVar
		self.legendTkVar.set(LEGEND)
		self.options1Sep = ttk.Separator(master = self.allOptionsFrame, orient = Tk.VERTICAL)
		self.options1Sep.pack(side = Tk.LEFT, padx = 2, expand = True, fill = Tk.BOTH)
		self.options2Frame = ttk.Frame(master = self.allOptionsFrame)
		self.options2Frame.pack(side = Tk.LEFT, padx = 5, pady = 5)
		for monomer in range(1, self.numMonomers + 1):
			colorFrame = ttk.Frame(master = self.options2Frame)
			colorFrame.pack(side = Tk.TOP, pady = 2)
			colorLabel = ttk.Label(master = colorFrame, text = "Monomer %s Color:  " %monomer)
			colorLabel.pack(side = Tk.LEFT)
			colorCanvas = Tk.Canvas(master = colorFrame, width = 15, height = 15)
			colorCanvas.pack(side = Tk. LEFT)
			colorCanvas.create_rectangle(0, 0, 20, 20, fill = COLORARRAY[monomer - 1])
			colorButton = ttk.Button(master = colorFrame, text = "Change", width = 8,
			 command = lambda monomer = monomer, canvas = colorCanvas:self.displayColorChooser(monomer, canvas))
			colorButton.pack(side = Tk.LEFT, padx = 5)
		for monomer in range(1, self.numMonomers + 1):
			dyadcolorFrame = ttk.Frame(master = self.options2Frame)
			dyadcolorFrame.pack(side = Tk.TOP, pady = 2)
			dyadcolorLabel = ttk.Label(master = dyadcolorFrame, text = "Homodyad %s Color:" %monomer)
			dyadcolorLabel.pack(side = Tk.LEFT)
			dyadcolorCanvas = Tk.Canvas(master = dyadcolorFrame, width = 15, height = 15)
			dyadcolorCanvas.pack(side = Tk. LEFT)
			dyadcolorCanvas.create_rectangle(0, 0, 20, 20, fill = DYADCOLORARRAY[monomer - 1 + self.numMonomers])
			dyadcolorButton = ttk.Button(master = dyadcolorFrame, text = "Change", width = 8,
			 command = lambda monomer = monomer, canvas = colorCanvas:self.displayDyadColorChooser(monomer, canvas))
			dyadcolorButton.pack(side = Tk.LEFT, padx = 5)
		self.options2Sep = ttk.Separator(master = self.allOptionsFrame, orient = Tk.VERTICAL)
		self.options2Sep.pack(side = Tk.LEFT, padx = 2, expand = True, fill = Tk.BOTH)
		self.options3Frame = ttk.Frame(master = self.allOptionsFrame)
		self.options3Frame.pack(side = Tk.LEFT, padx = 5)
		self.aliasTkVarList = []
		self.hphobTkVarList = []
		for monomer in range(1, self.numMonomers + 1):
			hydrophobFrame = ttk.Frame(master = self.options3Frame)
			hydrophobFrame.pack(side = Tk.TOP, pady = 3)
			hydrophobLabel = ttk.Label(master = hydrophobFrame, text = "Monomer %i: " %monomer)
			hydrophobLabel.pack(side = Tk.LEFT)
			hydrophobTkVar = Tk.StringVar()
			if self.hphobList:
				hydrophobTkVar.set(self.hphobList[monomer - 1])
			else:
				if HYDROPHOBICITY == 0:
					hydrophobTkVar.set("Hydrophobic")
				elif HYDROPHOBICITY == 1:
					hydrophobTkVar.set("Hydrophilic")
				else:
					hydrophobTkVar.set("None")
			hydrophobCombobox = ttk.Combobox(master = hydrophobFrame, values = ("Hydrophobic", "Hydrophilic", "None"),
			 textvariable = hydrophobTkVar, state = "readonly", width = 12)
			self.hphobTkVarList.append(hydrophobTkVar)
			hydrophobCombobox.pack(side = Tk.LEFT)
		for monomer in range(1, self.numMonomers + 1):
			aliasFrame = ttk.Frame(master = self.options3Frame)
			aliasFrame.pack(side = Tk.TOP, pady = 3)
			aliasLabel = ttk.Label(master = aliasFrame, text = "Monomer %i Alias:" %monomer)
			aliasLabel.pack(side = Tk.LEFT)	
			aliasTkVar = Tk.StringVar()
			if self.aliasList:
				aliasTkVar.set(self.aliasList[monomer - 1])
			aliasEntry = ttk.Entry(master = aliasFrame, width = 6, textvariable = aliasTkVar)
			self.aliasTkVarList.append(aliasTkVar)
			aliasEntry.pack(side = Tk.LEFT)
		self.okButton = ttk.Button(master = self.optionsWindow, text = "Apply", width = 14, command = lambda:self.apply())
		self.okButton.pack(side = Tk.TOP, pady = 3)
	def displayColorChooser(self, monomerID, canvas):
		color = askcolor()
		self.changeColor(color, monomerID)
		canvas.create_rectangle(0, 0, 20, 20, fill = color[1])
	def displayDyadColorChooser(self, monomerID, canvas):
		color = askcolor()
		self.changeDyadColor(color, monomerID)
		canvas.create_rectangle(0, 0, 20, 20, fill = color[1])
	def changeColor(self, color, monomerID):
		if color[1]:
			global COLORARRAY
			COLORARRAY[monomerID - 1] = color[1]
			print("color: ", color[1])
	def changeDyadColor(self, color, monomerID):
		if color[1]:
			global DYADCOLORARRAY
			DYADCOLORARRAY[monomerID - 1] = color[1]
	def apply(self):
		global DYAD
		DYAD = self.dyadTkVar.get()
		global LEGEND
		LEGEND = int(self.legendTkVar.get())
		global NUM_SIMULATIONS
		NUM_SIMULATIONS = int(self.numSimsTkVar.get())
		global MAINTAIN
		MAINTAIN = int(self.maintainTkVar.get())
		global STYLE
		STYLE = str(self.styleTkVar.get())
		self.newAliasList = []
		self.newHphobList = []
		applyAlias = False
		applyHphob = False
		for tkVar in self.aliasTkVarList:
			self.newAliasList.append(tkVar.get())
		if self.aliasList != self.newAliasList:
			self.aliasList = self.newAliasList
			applyAlias = True
		for tkVar in self.hphobTkVarList:
			self.newHphobList.append(tkVar.get())
		if self.hphobList != self.newHphobList:
			self.hphobList = self.newHphobList
			applyHphob = True	
		self.optionsWindow.destroy()
		if not applyAlias:
			return
		for alias in self.aliasList:
			if alias == "":
				global ALIAS
				ALIAS = False
				return
				print("reached here!")
				return
		ratioLabelCount = 0
		for ratioLabel in self.ratiosLabelList:
			ratioLabel.configure(text = self.aliasList[ratioLabelCount] + " Ratio:")
			ratioLabelCount += 1
		compLabelCount = 0
		for compLabel in self.compLabelList:
			compLabel.configure(text = self.aliasList[compLabelCount] + " Composition: ")
			compLabelCount += 1
		coeffLabelCount = 0
		for column in self.coeffLabelList:
			combinations = 0
			for coeffLabel in column:
				coeffLabel.configure(text = self.aliasList[coeffLabelCount] + "-" + self.aliasList[combinations] + " Constant:")
				coeffLabelCount += 1
				combinations += 1
			coeffLabelCount = 0
		global ALIAS
		ALIAS = True
		#self.amountFrame.destroy()
		#self.coefficientFrame.destroy()
		#self.createIterativeInputs(True)

	"----------------------------------------------------------------------------------------------------------------------------"
	"***DATA EXPORTATION***"
	"----------------------------------------------------------------------------------------------------------------------------"
	
	def export(self):
		if not self.plotted:
			errorMessage("Please simulate before exporting data!" , 330)
			return
		self.histDataList = []
		class histData:
			def __init__(self, name, data):
				self.name =  name
				self.data = data

		for polymerID in range(1, self.numMonomers + 1):
			data = self.getHistPlotvals(polymerID)[1]
			name = "Monomer %i Run Length" %(polymerID)
			currHistData = histData(name, data)
			self.histDataList.append(currHistData)
		wb = Workbook()
		ws = wb.active
		ws.title = "Run Length"
		colCount = 1
		maxRow = 0
		for histData in self.histDataList:
			ws.cell(row = 1, column = colCount, value = histData.name)
			ws.column_dimensions[get_column_letter(colCount)].width = len(histData.name) - 1
			ws.cell(row = 1, column = colCount + 1, value = "Normalized Counts")
			ws.column_dimensions[get_column_letter(colCount +1)].width = 20
			columnCount = 1
			maxRow = max(len(histData.data), maxRow)
			for column in ws.iter_cols(min_row = 2, max_row = len(histData.data) + 1, min_col = colCount, max_col = colCount):
				for cell in column:
					cell.value = columnCount
					columnCount += 1
			dataCount = 0
			for column in ws.iter_cols(min_row = 2, max_row = len(histData.data) + 1,min_col = colCount + 1, max_col = colCount + 1):
				for cell in column:	
					cell.value = histData.data[dataCount]
					dataCount += 1
			colCount += 2
		startRow = maxRow + 3
		ws2 = wb.create_sheet("Monomer Occurences")
		if DYAD:
			polymerArrayToUse = self.getDyad(self.polymerArray)
			monomerRange = self.numMonomers * 2 + 1
		else:
			polymerArrayToUse = self.polymerArray
			monomerRange = self.numMonomers + 1
		#Iterates through each unique monomer and adds it to a list of data
		monomerCountsList = []
		totalOriginalMonomerAmounts = sum(list(self.originalMonomerAmounts))
		conversionIndex = range(len(self.monomerUsageList[0]))
		conversionIndex = [i*self.numPolymers/totalOriginalMonomerAmounts*100 for i in conversionIndex]
		colCount = 1
		ws2.cell(row = 1, column = colCount, value = "Conversion Percentage")
		if not DYAD:
			for monomerID in range(1, self.numMonomers + 1):
				ws2.cell(row = 1, column = colCount + 1, value = "Normalized Monomer %i Occurence" %(monomerID))
				#ws.column_dimensions[get_column_letter(colCount +1)].width = 20
				colCount += 1
		if DYAD:
			for monomerID in range(1, 2*self.numMonomers + 1):
				ws2.cell(row = 1, column = colCount + 1, value = "Normalized Dyad %i Occurence" %(monomerID))
				#ws.column_dimensions[get_column_letter(colCount +1)].width = 20
				colCount += 1
		colCount = 1
		indexCount = 0
		if not DYAD:
			for column in ws2.iter_cols(min_row = 2, max_row = len(conversionIndex) + 1, min_col = colCount, max_col = colCount):
				for cell in column:
					cell.value = conversionIndex[indexCount]
					indexCount += 1
			for monomerID in range(1, self.numMonomers + 1):
				monomerIDcount = 0
				for column in ws2.iter_cols(min_row = 2, max_row = len(conversionIndex) + 1, min_col = colCount + 1, max_col = colCount + 1):
					for cell in column:
						cell.value = self.monomerUsageList[monomerID - 1][monomerIDcount]
						monomerIDcount += 1
				colCount += 1
		#DYAD IS DEPRECATED
		if DYAD:
			for column in ws2.iter_cols(min_row = 2, max_row = len(polymerIndex) + 1, min_col = colCount, max_col = colCount):
				for cell in column:
					cell.value = polymerIndex[indexCount]
					indexCount += 1
			for monomerID in range(1, 2*self.numMonomers + 1):
				monomerIDcount = 0
				for column in ws2.iter_cols(min_row = 2, max_row = len(polymerIndex) + 1, min_col = colCount + 1, max_col = colCount + 1):
					for cell in column:
						cell.value = monomerCountsList[monomerID - 1][monomerIDcount]
						monomerIDcount += 1
				colCount += 1
		ws3 = wb.create_sheet("Monomer Usage")
		for monomer in range(1, self.numMonomers + 1):
			#x-axis array
			polymerIndex = list(range(1, self.polymerLength + 1))
			#y-axis array initation
			monomercounts = [0] * self.polymerLength
			#variable to keep track of average number of monomers consumed
			monomersConsumed = 0
			#list of data for each monomer
			monomerRemaining = self.monomerRemainingList[0]
			conversionIndex = range(len(monomerRemaining))
			totalOriginalMonomerAmounts = sum(list(self.originalMonomerAmounts))
			conversionIndex = [i/totalOriginalMonomerAmounts*100 for i in conversionIndex]
		colCount = 1
		ws3.cell(row = 1, column = colCount, value = "Monomer Position Index")
		for monomerID in range(1, self.numMonomers + 1):
			ws3.cell(row = 1, column = colCount + 1, value = "%" + " of monomer %i remaining" %(monomerID))
			#ws.column_dimensions[get_column_letter(colCount +1)].width = 20
			colCount += 1
		colCount = 1
		indexCount = 0
		for column in ws3.iter_cols(min_row = 2, max_row = len(conversionIndex) + 1, min_col = colCount, max_col = colCount):
			for cell in column:
				cell.value = conversionIndex[indexCount]
				indexCount += 1
		for monomerID in range(1, self.numMonomers + 1):
			monomerIDcount = 0
			for column in ws3.iter_cols(min_row = 2, max_row = len(conversionIndex) + 1, min_col = colCount + 1, max_col = colCount + 1):
				for cell in column:
					cell.value = self.monomerRemainingList[monomerID - 1][monomerIDcount]
					monomerIDcount += 1
			colCount += 1
		# #x-axis array
		# polymerIndex = list(range(1, self.polymerLength + 1))
		# #list of data for all monomers
		# fullCompList = self.getFullCompositionAtIndex(self.polymerArray)
		# ws4 = wb.create_sheet("Polymer Compositions")
		# colCount = 1
		# ws4.cell(row = 1, column = colCount, value = "Monomer Position Index")
		# for monomerID in range(1, self.numMonomers + 1):
		# 	ws4.cell(row = 1, column = colCount + 1, value = "%" + "  Composition of Monomer %i" %(monomerID))
		# 	#ws.column_dimensions[get_column_letter(colCount +1)].width = 20
		# 	colCount += 1
		# colCount = 1
		# indexCount = 0
		# for column in ws4.iter_cols(min_row = 2, max_row = len(polymerIndex) + 1, min_col = colCount, max_col = colCount):
		# 	for cell in column:
		# 		cell.value = polymerIndex[indexCount]
		# 		indexCount += 1
		# for monomerID in range(1, self.numMonomers + 1):
		# 	monomerIDcount = 0
		# 	for column in ws4.iter_cols(min_row = 2, max_row = len(polymerIndex) + 1, min_col = colCount + 1, max_col = colCount + 1):
		# 		for cell in column:
		# 			cell.value = fullCompList[monomerID - 1][monomerIDcount]
		# 			monomerIDcount += 1
		# 	colCount += 1
		ws5 = wb.create_sheet("Hydrophobicity Blocks")
		class histData:
			def __init__(self, name, data):
				self.name =  name
				self.data = data
		self.histDataList = []
		polymerArray = self.convertHydro(self.polymerArray)
		histogramNumberLimit = int(self.histogramLimit * self.polymerLength)
		binwidth = 1
		for phobicity in range(0, 2):
			data, bins = np.histogram(self.getHistogramData(polymerArray, phobicity, histogramNumberLimit),
				bins=range(1, max(self.getHistogramData(polymerArray, phobicity, histogramNumberLimit)) + binwidth + 1, binwidth))
			#print(self.getHistogramData(self.polymerArray, polymerID, histogramNumberLimit))
			#print("data: ", data)
			#print('sum: ', sum(data))
			#print("divider: ", self.totalMonomers*self.histogramLimit)
			#print('length: ', len(self.getHistogramData(self.polymerArray, polymerID, histogramNumberLimit)))
			data = data / (self.polymerLength * self.numPolymers)
			if phobicity == 0:
				name = "Hydrophobic Blocks"
			elif phobicity == 1:
				name = "Hydrophilic Blocks"
			currHistData = histData(name, data)
			self.histDataList.append(currHistData)
			print("phobicity %i data: " %(phobicity), data)
		colCount = 1
		#indexCount = 0
		for histData in self.histDataList:
			ws5.cell(row = 1, column = colCount, value = histData.name)
			ws5.column_dimensions[get_column_letter(colCount)].width = len(histData.name) - 1
			ws5.cell(row = 1, column = colCount + 1, value = "Normalized Separation")
			ws5.column_dimensions[get_column_letter(colCount +1)].width = 20
			columnCount = 1
			maxRow = max(len(histData.data), maxRow)
			for column in ws5.iter_cols(min_row = 2, max_row = len(histData.data) + 1, min_col = colCount, max_col = colCount):
				for cell in column:
					cell.value = columnCount
					columnCount += 1
			dataCount = 0
			for column in ws5.iter_cols(min_row = 2, max_row = len(histData.data) + 1,min_col = colCount + 1, max_col = colCount + 1):
				for cell in column:	
					cell.value = histData.data[dataCount]
					dataCount += 1
			colCount += 2
		name = "graphData"
		try:
			wb.save(name + ".xlsx")
			infoMessage("Export Successful", "Data successfully exported to graphData.xlsx!", 330)
		except: 
			errorMessage("Cannot export if graphData.xlsx is open!", 300)
		return
		
"-------------------------------------------------------------------------------------------------------------------------------"
"***GLOBAL FUNCTIONS***"
"-------------------------------------------------------------------------------------------------------------------------------"

"**Error message function***"
#When called, makes a pop out error informing user of invalid inputs
def errorMessage(message, width):
	#Toplevel parameters
	top = Tk.Toplevel()
	top.grab_set()
	top.focus_force()
	#top.grab_set()
	top.wm_title("Error")
	top.geometry("%dx%d%+d%+d" % (width, 70, 250, 125))
	#Message
	msg = Tk.Message(master = top, text = message, width = 500)
	msg.pack(side = Tk.TOP, pady = 5)
	#OK button to exit
	exitButton = ttk.Button(master = top, text = "Ok", command = top.destroy, width = 7)
	exitButton.pack(side = Tk.TOP, pady = 5)

"***Info message function***"
def infoMessage(title, message, width):
	#Toplevel parameters
	top = Tk.Toplevel()
	top.grab_set()
	top.focus_force()
	top.wm_title(title)
	top.geometry("%dx%d%+d%+d" % (width, 70, 250, 125))
	#Message
	msg = Tk.Message(master = top, text = message, width = 500)
	msg.pack(side = Tk.TOP, pady = 5)
	#OK button to exit
	exitButton = ttk.Button(master = top, text = "Ok", command = top.destroy, width = 7)
	exitButton.pack(side = Tk.TOP, pady = 5)

"***Weighted choice function***"
#Takes in a list of tuple lists with item and weight, and returns a random item based on 
#weight
def weighted_choice(choices):
    total = sum(w for c, w in choices)
    r = random.uniform(0, total)
    upto = 0
    for c, w in choices:
        if upto + w >= r:
        	return c
        upto += w
    assert False, "Shouldn't get here"

"*** Window Centering function****"
def center(toplevel):
    toplevel.update_idletasks()
    w = toplevel.winfo_screenwidth()
    h = toplevel.winfo_screenheight()
    size = tuple(int(_) for _ in toplevel.geometry().split('+')[0].split('x'))
    x = w/2 - size[0]/2
    y = h/2 - size[1]/2
    toplevel.geometry("%dx%d+%d+%d" % (size + (x, y - 30)))
#def flatten(l):
#	flat_list = [item for sublist in l for item in sublist]
#	return flat_list
"***Proper shutdown function***"
def on_closing():
	root.quit()
	root.destroy()

"--------------------------------------------------------------------------------------------------------------------------"
"***USER DEFINED ERRORS***"
"--------------------------------------------------------------------------------------------------------------------------"

class notInEuropeError(Exception):
	def __init__(self, value):
		self.value = value

class phobicityNotSpecified(Exception):
	def __init__(self, value):
		self.value = value

"-----------------------------------------------------------------------------------------------------------------------------"
"***STARTS THE APPLICATION***"
"-----------------------------------------------------------------------------------------------------------------------------"

root = Tk.Tk()
root.protocol("WM_DELETE_WINDOW", on_closing)
root.wm_title("Compositional Drift %s" % VERSION)
app = Application(master = root)
app.mainloop()
